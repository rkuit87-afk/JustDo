"""
reports.py  -  V8
Generates Excel reports aligned to the Equipment Availability & Reliability template.
"""
import os, smtplib, sqlite3, logging
from datetime import datetime, date, timedelta
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from database import get_db
from scheduler import get_week_start

logger = logging.getLogger(__name__)

try:
    from config import ARCHIVE_MONTHS, SITE_NAME
except ImportError:
    ARCHIVE_MONTHS = 6
    SITE_NAME = 'Mill Maintenance Terminal'

def _get_mail_settings():
    from database import get_setting
    return (
        get_setting('mail_sender', ''),
        get_setting('mail_password', ''),
        get_setting('mail_recipient', '')
    )

def _get_recipients(key):
    from database import get_setting
    import json
    try:
        return json.loads(get_setting(key, '[]'))
    except (ValueError, TypeError) as e:
        logger.warning(f"Failed to parse recipients for '{key}': {e}")
        return []

REPORTS_DIR = os.path.join(os.path.dirname(__file__), 'reports')
os.makedirs(REPORTS_DIR, exist_ok=True)

# -- Styles --
HDR_FILL = PatternFill("solid", fgColor="1F3864")   # dark blue header
HDR_FONT = Font(color="FFFFFF", bold=True, size=10)
SUB_FILL = PatternFill("solid", fgColor="D6DCE4")
SUB_FONT = Font(bold=True, size=10)
RED_FILL = PatternFill("solid", fgColor="FCE4D6")
GRN_FILL = PatternFill("solid", fgColor="E2EFDA")
YLW_FILL = PatternFill("solid", fgColor="FFEB9C")
BLU_FILL = PatternFill("solid", fgColor="DDEBF7")   # variable input cells (green equiv)
THIN     = Border(
    left=Side(style='thin', color='BFBFBF'),
    right=Side(style='thin', color='BFBFBF'),
    top=Side(style='thin', color='BFBFBF'),
    bottom=Side(style='thin', color='BFBFBF')
)

def style_header(ws, row, cols):
    for col in range(1, cols+1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN

def style_row(ws, row, cols, fill=None):
    for col in range(1, cols+1):
        cell = ws.cell(row=row, column=col)
        if fill: cell.fill = fill
        cell.border = THIN
        cell.alignment = Alignment(vertical='top', wrap_text=True)

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def write_title(ws, title, subtitle=None):
    ws['A1'] = title
    ws['A1'].font = Font(bold=True, size=14, color='1F3864')
    if subtitle:
        ws['A2'] = subtitle
        ws['A2'].font = Font(size=10, color='808080')
    ws.sheet_view.showGridLines = False


# ============================================================
# WEEKLY REPORT
# ============================================================
def generate_weekly_report(week_start=None):
    week_start = week_start or get_week_start()
    week_end   = week_start + timedelta(days=6)
    conn = get_db()
    c    = conn.cursor()
    wb   = Workbook()

    # ---- Sheet 1: Summary ----
    ws = wb.active
    ws.title = "Weekly Summary"
    write_title(ws,
        f"Weekly Maintenance Report - {get_setting_local('site_name', SITE_NAME)}",
        f"Week: {week_start.strftime('%d %b %Y')} to {week_end.strftime('%d %b %Y')}"
    )
    ws.row_dimensions[3].height = 8

    # KPI boxes row
    open_bds  = c.execute("SELECT COUNT(*) FROM breakdowns WHERE status='supervisor_logged'").fetchone()[0]
    disputed  = c.execute("SELECT COUNT(*) FROM breakdowns WHERE status='disputed'").fetchone()[0]
    total_bd  = c.execute("""
        SELECT COUNT(*), COALESCE(SUM(downtime_mins),0)
        FROM breakdowns WHERE DATE(logged_at) BETWEEN ? AND ?
    """, (week_start.isoformat(), week_end.isoformat())).fetchone()

    ws['A4'] = 'Total Breakdowns'; ws['A4'].font = Font(bold=True, size=11)
    ws['B4'] = total_bd[0];        ws['B4'].font = Font(bold=True, size=16, color='C00000')
    ws['C4'] = 'Total Downtime (hrs)'; ws['C4'].font = Font(bold=True, size=11)
    ws['D4'] = round((total_bd[1] or 0)/60, 1); ws['D4'].font = Font(bold=True, size=16, color='C00000')
    ws['E4'] = 'Pending Artisan';  ws['E4'].font = Font(bold=True, size=11)
    ws['F4'] = open_bds;           ws['F4'].font = Font(bold=True, size=16, color='ED7D31')
    ws['G4'] = 'Disputed';         ws['G4'].font = Font(bold=True, size=11)
    ws['H4'] = disputed;           ws['H4'].font = Font(bold=True, size=16, color='FF0000')
    ws.row_dimensions[5].height = 8

    # Machine downtime table
    row = 6
    ws.cell(row=row, column=1, value='MACHINE DOWNTIME THIS WEEK').font = SUB_FONT
    ws.cell(row=row, column=1).fill = SUB_FILL
    row += 1

    bd_headers = ['Equipment', 'Area / Section', 'Failure Count',
                  'Downtime (hrs)', 'Downtime (min)', 'Maint. Down (min)',
                  'Prod. Down (min)', 'Failure Type', 'Components']
    for i, h in enumerate(bd_headers, 1):
        ws.cell(row=row, column=i, value=h)
    style_header(ws, row, len(bd_headers))
    ws.row_dimensions[row].height = 30
    row += 1

    bds = c.execute("""
        SELECT e.name as equip_name, e.area, e.section,
               COUNT(*) as failures,
               COALESCE(SUM(b.downtime_mins),0) as total_mins,
               COALESCE(SUM(b.maintenance_downtime_mins),0) as maint_mins,
               COALESCE(SUM(b.production_downtime_mins),0) as prod_mins,
               GROUP_CONCAT(DISTINCT COALESCE(b.final_failure_type, b.sup_failure_type)) as ftypes,
               GROUP_CONCAT(DISTINCT b.sup_component) as components
        FROM   breakdowns b
        JOIN   equipment e ON e.id = b.equipment_id
        WHERE  DATE(b.logged_at) BETWEEN ? AND ?
        GROUP  BY e.id ORDER BY total_mins DESC
    """, (week_start.isoformat(), week_end.isoformat())).fetchall()

    total_mins = 0
    for bd in bds:
        mins = bd['total_mins'] or 0
        total_mins += mins
        fill = RED_FILL if mins > 120 else YLW_FILL if mins > 30 else None
        vals = [
            bd['equip_name'],
            f"{bd['area']} / {bd['section'] or '-'}",
            bd['failures'],
            round(mins/60, 2),
            mins,
            bd['maint_mins'] or 0,
            bd['prod_mins'] or 0,
            bd['ftypes'] or '',
            bd['components'] or ''
        ]
        for i, v in enumerate(vals, 1):
            ws.cell(row=row, column=i, value=v)
        style_row(ws, row, len(bd_headers), fill)
        row += 1

    # Total row
    ws.cell(row=row, column=1, value='TOTAL').font = Font(bold=True)
    ws.cell(row=row, column=4, value=round(total_mins/60,2)).font = Font(bold=True)
    ws.cell(row=row, column=5, value=total_mins).font = Font(bold=True)
    style_row(ws, row, len(bd_headers), SUB_FILL)
    row += 2

    # PM Compliance
    ws.cell(row=row, column=1, value='PM COMPLIANCE THIS WEEK').font = SUB_FONT
    ws.cell(row=row, column=1).fill = SUB_FILL
    row += 1

    pm_hdr = ['Artisan', 'Scheduled', 'Completed', 'Missed', 'Compliance %']
    for i, h in enumerate(pm_hdr, 1):
        ws.cell(row=row, column=i, value=h)
    style_header(ws, row, len(pm_hdr))
    row += 1

    compliance = c.execute("""
        SELECT u.name, COUNT(*) as total,
               SUM(CASE WHEN t.status='done' THEN 1 ELSE 0 END) as done,
               SUM(CASE WHEN t.status='missed' THEN 1 ELSE 0 END) as missed
        FROM   pm_tasks t JOIN users u ON u.id = t.assigned_to
        WHERE  t.week_start=?
        GROUP  BY u.name
    """, (week_start.isoformat(),)).fetchall()

    for pm in compliance:
        pct  = round((pm['done']/pm['total']*100) if pm['total'] else 0, 1)
        fill = GRN_FILL if pct >= 90 else YLW_FILL if pct >= 70 else RED_FILL
        vals = [pm['name'], pm['total'], pm['done'], pm['missed'] or 0, f"{pct}%"]
        for i, v in enumerate(vals, 1):
            ws.cell(row=row, column=i, value=v)
        style_row(ws, row, len(pm_hdr), fill)
        row += 1

    set_col_widths(ws, [32, 18, 12, 12, 12, 14, 14, 18, 30])

    # ---- Sheet 2: Daily Downtime Log (matches template) ----
    ws2 = wb.create_sheet("Daily Downtime Log")
    ws2.sheet_view.showGridLines = False
    write_title(ws2,
        "Daily Downtime Log",
        f"{week_start.strftime('%d %b %Y')} to {week_end.strftime('%d %b %Y')}"
    )

    dl_headers = [
        'Date', 'Production Dept', 'Equipment ID', 'Equipment Name',
        'Failure Description', 'Downtime Start', 'Downtime End',
        'Downtime Hours', 'Downtime Minutes',
        'Failure Type', 'Maint. Downtime Hrs', 'Failed Component',
        'Failure Mode', 'Equip ID (Maint)', 'Prod. Downtime Hrs',
        'Equip ID (Prod)', 'Repair Time', 'Comments'
    ]
    row = 4
    for i, h in enumerate(dl_headers, 1):
        ws2.cell(row=row, column=i, value=h)
    style_header(ws2, row, len(dl_headers))
    ws2.row_dimensions[row].height = 30
    row += 1

    bds_detail = c.execute("""
        SELECT b.*, e.name as equip_name, e.area, e.equipment_id as equip_code,
               ua.name as artisan_name, us.name as supervisor_name,
               GROUP_CONCAT(bp.part_desc || ' x' || bp.qty, ', ') as parts_list
        FROM   breakdowns b
        JOIN   equipment e  ON e.id  = b.equipment_id
        LEFT JOIN users ua  ON ua.id = b.artisan_id
        LEFT JOIN users us  ON us.id = b.supervisor_id
        LEFT JOIN breakdown_parts bp ON bp.breakdown_id = b.id
        WHERE  DATE(b.logged_at) BETWEEN ? AND ?
        GROUP  BY b.id ORDER BY b.logged_at ASC
    """, (week_start.isoformat(), week_end.isoformat())).fetchall()

    for bd in bds_detail:
        t_start  = bd['final_time_start'] or bd['sup_time_start'] or ''
        t_end    = bd['final_time_end']   or bd['sup_time_end']   or ''
        d_mins   = bd['downtime_mins'] or 0
        d_hrs    = round(d_mins/60, 4) if d_mins else 0
        maint_hrs= round((bd['maintenance_downtime_mins'] or 0)/60, 4)
        prod_hrs = round((bd['production_downtime_mins']  or 0)/60, 4)
        ftype    = bd['final_failure_type'] or bd['sup_failure_type'] or ''
        # For template: maint equip ID if maintenance downtime, prod equip ID if production
        maint_equip = bd['equip_code'] if maint_hrs > 0 else ''
        prod_equip  = bd['equip_code'] if prod_hrs  > 0 else ''

        vals = [
            bd['log_date'] or bd['logged_at'][:10],
            bd['area'],
            bd['equip_code'],
            bd['equip_name'],
            bd['art_description'] or bd['sup_notes'] or '',
            t_start,
            t_end,
            d_hrs,
            d_mins,
            ftype,
            maint_hrs,
            bd['sup_component'] or '',
            bd['sup_failure_mode'] or '',
            maint_equip,
            prod_hrs,
            prod_equip,
            '',  # Repair time - blank, can be filled manually
            bd['art_repair_action'] or ''
        ]
        for i, v in enumerate(vals, 1):
            ws2.cell(row=row, column=i, value=v or '')
        # Colour: green = variable input, yellow = auto-calculated
        for col in [1,2,3,4,5,6,7]:   # variable inputs (green)
            ws2.cell(row=row, column=col).fill = GRN_FILL
        for col in [8,9,11,15]:        # auto-calculated (yellow)
            ws2.cell(row=row, column=col).fill = YLW_FILL
        style_row(ws2, row, len(dl_headers))
        row += 1

    set_col_widths(ws2, [11,12,12,28,35,8,8,10,10,14,12,20,18,12,12,12,10,30])

    # ---- Sheet 3: Breakdowns Detail ----
    ws3 = wb.create_sheet("Breakdowns Detail")
    ws3.sheet_view.showGridLines = False
    write_title(ws3, "Breakdown Log Detail",
                f"{week_start.strftime('%d %b')} to {week_end.strftime('%d %b %Y')}")

    bd3_headers = ['Date/Time', 'Equipment', 'Area', 'Artisan', 'Supervisor',
                   'Start', 'Restored', 'Downtime (min)', 'Failure Type (Final)',
                   'Component', 'Failure Mode', 'Description', 'Repair Action',
                   'Machine Status', 'Follow-up', 'Parts Used', 'Status', 'Disputed']
    row = 4
    for i, h in enumerate(bd3_headers, 1):
        ws3.cell(row=row, column=i, value=h)
    style_header(ws3, row, len(bd3_headers))
    row += 1

    for bd in bds_detail:
        fill = RED_FILL if (bd['downtime_mins'] or 0) > 120 else None
        vals = [
            bd['logged_at'], bd['equip_name'], bd['area'],
            bd['artisan_name'], bd['supervisor_name'],
            bd['final_time_start'] or bd['sup_time_start'],
            bd['final_time_end']   or bd['sup_time_end'],
            bd['downtime_mins'],
            bd['final_failure_type'] or bd['sup_failure_type'],
            bd['sup_component'], bd['sup_failure_mode'],
            bd['art_description'], bd['art_repair_action'],
            bd['art_machine_status'] or bd['sup_machine_status'],
            bd['art_followup'], bd['parts_list'],
            bd['status'], 'Yes' if bd['disputed'] else 'No'
        ]
        for i, v in enumerate(vals, 1):
            ws3.cell(row=row, column=i, value=v or '')
        style_row(ws3, row, len(bd3_headers), fill)
        row += 1

    set_col_widths(ws3, [16,30,10,16,16,7,7,12,16,18,16,35,25,25,10,30,12,8])

    # ---- Sheet 4: PM Tasks ----
    ws4 = wb.create_sheet("PM Tasks")
    ws4.sheet_view.showGridLines = False
    write_title(ws4, "Planned Maintenance",
                f"{week_start.strftime('%d %b')} to {week_end.strftime('%d %b %Y')}")

    pm_headers = ['Artisan', 'Task', 'Equipment', 'Area', 'Shift',
                  'Status', 'Completed At', 'Outcome', 'Condition', 'Notes', 'Parts Used']
    row = 4
    for i, h in enumerate(pm_headers, 1):
        ws4.cell(row=row, column=i, value=h)
    style_header(ws4, row, len(pm_headers))
    row += 1

    pm_tasks = c.execute("""
        SELECT t.*, u.name as artisan_name, e.name as equip_name, e.area,
               uc.name as completed_by_name, s.shift,
               GROUP_CONCAT(pp.part_desc || ' x' || pp.qty, ', ') as parts
        FROM   pm_tasks t
        JOIN   users u    ON u.id  = t.assigned_to
        LEFT JOIN equipment e ON e.id = t.equipment_id
        LEFT JOIN users uc ON uc.id = t.completed_by
        LEFT JOIN pm_schedule s ON s.id = t.schedule_id
        LEFT JOIN pm_parts pp ON pp.task_id = t.id
        WHERE  t.week_start=?
        GROUP  BY t.id
        ORDER  BY u.name, t.status, t.title
    """, (week_start.isoformat(),)).fetchall()

    for t in pm_tasks:
        fill = GRN_FILL if t['status']=='done' else RED_FILL if t['status']=='missed' else None
        vals = [
            t['artisan_name'], t['title'], t['equip_name'], t['area'],
            t['shift'], t['status'].upper(), t['completed_at'],
            t['outcome'], t['condition_found'], t['notes'], t['parts']
        ]
        for i, v in enumerate(vals, 1):
            ws4.cell(row=row, column=i, value=v or '')
        style_row(ws4, row, len(pm_headers), fill)
        row += 1

    set_col_widths(ws4, [16,35,30,10,6,8,16,25,20,30,30])

    # ---- Sheet 5: Boiler Operations (matches template) ----
    ws5 = wb.create_sheet("Boiler Operations")
    ws5.sheet_view.showGridLines = False
    write_title(ws5, "Boiler Operations Log",
                f"{week_start.strftime('%d %b')} to {week_end.strftime('%d %b %Y')}")

    bo_headers = [
        'Date',
        'B1 TDS Morning', 'B1 TDS Night', 'B1 Highest', 'B1 Blowdowns',
        'B2 TDS Morning', 'B2 TDS Night', 'B2 Highest', 'B2 Blowdowns',
        'B3 TDS Morning', 'B3 TDS Night', 'B3 Highest', 'B3 Blowdowns',
        'TDS Setpoint',
        'Condensate Meter', 'Condensate Return',
        'Makeup Meter', 'Actual Makeup',
        'Softener Day', 'Softener Night', 'Salt Bags (50kg)',
        'Lost Time (min)', 'Downtime Incident', 'Operator Comments'
    ]
    row = 4
    for i, h in enumerate(bo_headers, 1):
        ws5.cell(row=row, column=i, value=h)
    style_header(ws5, row, len(bo_headers))
    ws5.row_dimensions[row].height = 45
    row += 1

    boilers = c.execute("""
        SELECT id, name FROM equipment WHERE area='Boilers' ORDER BY name
    """).fetchall()
    boiler_ids = [b['id'] for b in boilers]

    # Get readings grouped by date
    since = week_start.isoformat()
    until = week_end.isoformat()
    readings_raw = c.execute("""
        SELECT br.*, e.name as boiler_name
        FROM boiler_readings br
        JOIN equipment e ON e.id = br.boiler_id
        WHERE br.log_date BETWEEN ? AND ?
        ORDER BY br.log_date, e.name
    """, (since, until)).fetchall()

    # Group by date
    from collections import defaultdict
    by_date = defaultdict(dict)
    for r in readings_raw:
        by_date[r['log_date']][r['boiler_id']] = dict(r)

    for log_date in sorted(by_date.keys()):
        day_data = by_date[log_date]
        row_vals = [log_date]
        for b_id in boiler_ids[:3]:  # max 3 boilers
            rd = day_data.get(b_id, {})
            row_vals += [
                rd.get('tds_morning',''),
                rd.get('tds_night',''),
                rd.get('tds_highest',''),
                rd.get('blowdowns_total',0),
            ]
        # Pad if fewer than 3 boilers
        while len(row_vals) < 13:
            row_vals += ['','','',0]

        # Use first boiler's shared readings
        first = list(day_data.values())[0] if day_data else {}
        row_vals += [
            3300,  # TDS setpoint
            first.get('condensate_meter_reading',''),
            first.get('condensate_return',''),
            first.get('makeup_meter_reading',''),
            first.get('makeup_water',''),
            first.get('softener_day',''),
            first.get('softener_night',''),
            first.get('salt_bags',0),
            first.get('lost_time_mins',0),
            first.get('downtime_incident',''),
            first.get('operator_comments',''),
        ]
        for i, v in enumerate(row_vals, 1):
            ws5.cell(row=row, column=i, value=v)
        # Variable input cols = green, auto-calc = yellow
        for col in [2,3,6,7,10,11,15,17,19,20,21,22,23,24]:  # inputs
            ws5.cell(row=row, column=col).fill = GRN_FILL
        for col in [4,5,8,9,12,13,16,18]:  # calculated
            ws5.cell(row=row, column=col).fill = YLW_FILL
        style_row(ws5, row, len(bo_headers))
        row += 1

    set_col_widths(ws5, [11,10,10,10,10,10,10,10,10,10,10,10,10,10,14,14,14,14,12,12,10,12,25,30])

    # ---- Sheet 6: Generator Runtime ----
    ws6 = wb.create_sheet("Generator Runtime")
    ws6.sheet_view.showGridLines = False
    write_title(ws6, "Generator Runtime Log",
                f"{week_start.strftime('%d %b')} to {week_end.strftime('%d %b %Y')}")

    generators = c.execute("""
        SELECT id, name, equipment_id FROM equipment
        WHERE area='Utility' AND section='Generators' ORDER BY name
    """).fetchall()

    # Build headers dynamically per generator
    gen_headers = ['Date']
    for g in generators:
        gen_headers += [f"{g['name']} - Hour Meter", f"{g['name']} - Run Hrs", f"{g['name']} - Diesel Issued (L)"]
    gen_headers.append('Combined Avg Run Time (hrs)')

    row = 4
    for i, h in enumerate(gen_headers, 1):
        ws6.cell(row=row, column=i, value=h)
    style_header(ws6, row, len(gen_headers))
    ws6.row_dimensions[row].height = 45
    row += 1

    gen_readings = c.execute("""
        SELECT ur.*, e.name as equip_name
        FROM utility_readings ur
        JOIN equipment e ON e.id = ur.substation_id
        WHERE e.section='Generators' AND ur.log_date BETWEEN ? AND ?
        ORDER BY ur.log_date, e.name
    """, (since, until)).fetchall()

    by_date_gen = defaultdict(dict)
    for r in gen_readings:
        by_date_gen[r['log_date']][r['equipment_id']] = dict(r)

    for log_date in sorted(by_date_gen.keys()):
        day_data = by_date_gen[log_date]
        row_vals = [log_date]
        run_hrs_list = []
        for g in generators:
            rd = day_data.get(g['id'], {})
            run_h = rd.get('run_hours_today') or 0
            run_hrs_list.append(run_h)
            row_vals += [
                rd.get('hour_meter',''),
                run_h,
                rd.get('diesel_issued',0),
            ]
        avg = round(sum(run_hrs_list)/len(run_hrs_list), 2) if run_hrs_list else 0
        row_vals.append(avg)
        for i, v in enumerate(row_vals, 1):
            ws6.cell(row=row, column=i, value=v)
        style_row(ws6, row, len(gen_headers))
        row += 1

    set_col_widths(ws6, [11] + [14,10,12]*len(generators) + [16])

    # ---- Sheet 7: Compressor Runtime ----
    ws7 = wb.create_sheet("Compressor Runtime")
    ws7.sheet_view.showGridLines = False
    write_title(ws7, "Compressor Runtime Log",
                f"{week_start.strftime('%d %b')} to {week_end.strftime('%d %b %Y')}")

    compressors = c.execute("""
        SELECT id, name, equipment_id FROM equipment
        WHERE area='Utility' AND section='Compressed Air' ORDER BY name
    """).fetchall()

    comp_headers = ['Date']
    for comp in compressors:
        comp_headers += [f"{comp['name']} - Hour Meter", f"{comp['name']} - Run Hrs"]

    row = 4
    for i, h in enumerate(comp_headers, 1):
        ws7.cell(row=row, column=i, value=h)
    style_header(ws7, row, len(comp_headers))
    ws7.row_dimensions[row].height = 45
    row += 1

    comp_readings = c.execute("""
        SELECT ur.*, e.name as equip_name
        FROM utility_readings ur
        JOIN equipment e ON e.id = ur.substation_id
        WHERE e.section='Compressed Air' AND ur.log_date BETWEEN ? AND ?
        ORDER BY ur.log_date, e.name
    """, (since, until)).fetchall()

    by_date_comp = defaultdict(dict)
    for r in comp_readings:
        by_date_comp[r['log_date']][r['equipment_id']] = dict(r)

    for log_date in sorted(by_date_comp.keys()):
        day_data = by_date_comp[log_date]
        row_vals = [log_date]
        for comp in compressors:
            rd = day_data.get(comp['id'], {})
            row_vals += [
                rd.get('hour_meter',''),
                rd.get('run_hours_today',0),
            ]
        for i, v in enumerate(row_vals, 1):
            ws7.cell(row=row, column=i, value=v)
        style_row(ws7, row, len(comp_headers))
        row += 1

    set_col_widths(ws7, [11] + [14,10]*len(compressors))

    # ---- Sheet 8: Spares Reorder ----
    ws8 = wb.create_sheet("Spares Reorder")
    ws8.sheet_view.showGridLines = False
    write_title(ws8, "Spares Reorder List")

    ro_headers = ['Part Description', 'Qty', 'Source', 'Date Logged']
    row = 4
    for i, h in enumerate(ro_headers, 1):
        ws8.cell(row=row, column=i, value=h)
    style_header(ws8, row, len(ro_headers))
    row += 1

    reorder = c.execute("""
        SELECT part_desc, SUM(qty) as qty,
               GROUP_CONCAT(DISTINCT source) as sources,
               MAX(logged_at) as last_logged
        FROM   reorder WHERE cleared=0
        GROUP  BY LOWER(TRIM(part_desc)) ORDER BY part_desc
    """).fetchall()

    for r in reorder:
        ws8.cell(row=row, column=1, value=r['part_desc'])
        ws8.cell(row=row, column=2, value=r['qty'])
        ws8.cell(row=row, column=3, value=r['sources'])
        ws8.cell(row=row, column=4, value=r['last_logged'])
        style_row(ws8, row, 4)
        row += 1

    set_col_widths(ws8, [40, 8, 35, 16])

    conn.close()
    fname = f"weekly_report_{week_start.strftime('%Y_%m_%d')}.xlsx"
    fpath = os.path.join(REPORTS_DIR, fname)
    wb.save(fpath)
    return fpath


# ============================================================
# MONTHLY ARCHIVE
# ============================================================
def generate_monthly_archive(year, month):
    from calendar import monthrange
    last_day = monthrange(year, month)[1]
    start    = date(year, month, 1).isoformat()
    end      = date(year, month, last_day).isoformat()
    conn     = get_db()
    c        = conn.cursor()
    wb       = Workbook()
    month_str = date(year, month, 1).strftime('%B %Y')
    site_name = get_setting_local('site_name', SITE_NAME)

    # ---- Breakdowns ----
    ws = wb.active
    ws.title = 'Breakdowns'
    write_title(ws, f'Breakdown Archive - {month_str}', site_name)

    headers = ['Date', 'Equipment', 'Area', 'Section', 'Artisan',
               'Downtime (min)', 'Maint. Down (min)', 'Prod. Down (min)',
               'Failure Type (Final)', 'Component', 'Failure Mode',
               'Description', 'Repair Action', 'Machine Status', 'Parts Used',
               'Status', 'Disputed', 'Resolution Note']
    row = 4
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
    style_header(ws, row, len(headers))
    row += 1

    bds = c.execute("""
        SELECT b.log_date, e.name, e.area, e.section, ua.name as artisan,
               b.downtime_mins, b.maintenance_downtime_mins, b.production_downtime_mins,
               COALESCE(b.final_failure_type, b.sup_failure_type) as ftype,
               b.sup_component, b.sup_failure_mode,
               b.art_description, b.art_repair_action,
               COALESCE(b.art_machine_status, b.sup_machine_status) as mstatus,
               GROUP_CONCAT(bp.part_desc || ' x' || bp.qty, ', ') as parts_list,
               b.status, b.disputed, b.admin_resolution_note
        FROM   breakdowns b
        JOIN   equipment e ON e.id = b.equipment_id
        LEFT JOIN users ua ON ua.id = b.artisan_id
        LEFT JOIN breakdown_parts bp ON bp.breakdown_id = b.id
        WHERE  DATE(b.logged_at) BETWEEN ? AND ?
        GROUP  BY b.id ORDER BY b.logged_at
    """, (start, end)).fetchall()

    for bd in bds:
        fill = RED_FILL if (bd['downtime_mins'] or 0) > 120 else None
        for i, v in enumerate(bd, 1):
            ws.cell(row=row, column=i, value=v or '')
        style_row(ws, row, len(headers), fill)
        row += 1

    set_col_widths(ws, [11,30,10,12,16,12,12,12,16,18,16,35,25,25,30,12,8,30])

    # ---- Boiler Operations ----
    ws2 = wb.create_sheet('Boiler Operations')
    write_title(ws2, f'Boiler Operations - {month_str}', site_name)

    bo_hdr = [
        'Date',
        'B1 TDS Morn', 'B1 TDS Night', 'B1 Highest', 'B1 Blowdowns',
        'B2 TDS Morn', 'B2 TDS Night', 'B2 Highest', 'B2 Blowdowns',
        'B3 TDS Morn', 'B3 TDS Night', 'B3 Highest', 'B3 Blowdowns',
        'Condensate Return', 'Actual Makeup',
        'Softener Day', 'Softener Night', 'Salt Bags',
        'Lost Time (min)', 'Downtime Incident', 'Comments'
    ]
    row = 4
    for i, h in enumerate(bo_hdr, 1):
        ws2.cell(row=row, column=i, value=h)
    style_header(ws2, row, len(bo_hdr))
    row += 1

    boilers = c.execute(
        "SELECT id FROM equipment WHERE area='Boilers' ORDER BY name"
    ).fetchall()
    boiler_ids = [b['id'] for b in boilers]

    readings_raw = c.execute("""
        SELECT br.* FROM boiler_readings br
        WHERE br.log_date BETWEEN ? AND ?
        ORDER BY br.log_date, br.boiler_id
    """, (start, end)).fetchall()

    from collections import defaultdict
    by_date = defaultdict(dict)
    for r in readings_raw:
        by_date[r['log_date']][r['boiler_id']] = dict(r)

    for log_date in sorted(by_date.keys()):
        day_data = by_date[log_date]
        row_vals = [log_date]
        for b_id in boiler_ids[:3]:
            rd = day_data.get(b_id, {})
            row_vals += [rd.get('tds_morning',''), rd.get('tds_night',''),
                         rd.get('tds_highest',''), rd.get('blowdowns_total',0)]
        while len(row_vals) < 13:
            row_vals += ['','','',0]
        first = list(day_data.values())[0] if day_data else {}
        row_vals += [
            first.get('condensate_return',''),
            first.get('makeup_water',''),
            first.get('softener_day',''),
            first.get('softener_night',''),
            first.get('salt_bags',0),
            first.get('lost_time_mins',0),
            first.get('downtime_incident',''),
            first.get('operator_comments',''),
        ]
        for i, v in enumerate(row_vals, 1):
            ws2.cell(row=row, column=i, value=v)
        style_row(ws2, row, len(bo_hdr))
        row += 1

    set_col_widths(ws2, [11,10,10,10,10,10,10,10,10,10,10,10,10,14,14,12,12,10,12,25,30])

    # ---- Generator Runtime ----
    ws3 = wb.create_sheet('Generator Runtime')
    write_title(ws3, f'Generator Runtime - {month_str}', site_name)

    generators = c.execute(
        "SELECT id, name FROM equipment WHERE area='Utility' AND section='Generators' ORDER BY name"
    ).fetchall()
    gen_hdr = ['Date']
    for g in generators:
        gen_hdr += [f"{g['name']} - Hour Meter", f"{g['name']} - Run Hrs", f"{g['name']} - Diesel (L)"]

    row = 4
    for i, h in enumerate(gen_hdr, 1):
        ws3.cell(row=row, column=i, value=h)
    style_header(ws3, row, len(gen_hdr))
    row += 1

    gen_readings = c.execute("""
        SELECT ur.* FROM utility_readings ur
        JOIN equipment e ON e.id = ur.substation_id
        WHERE e.section='Generators' AND ur.log_date BETWEEN ? AND ?
        ORDER BY ur.log_date, e.name
    """, (start, end)).fetchall()

    by_date_g = defaultdict(dict)
    for r in gen_readings:
        by_date_g[r['log_date']][r['equipment_id']] = dict(r)

    for log_date in sorted(by_date_g.keys()):
        day_data = by_date_g[log_date]
        row_vals = [log_date]
        for g in generators:
            rd = day_data.get(g['id'], {})
            row_vals += [rd.get('hour_meter',''), rd.get('run_hours_today',0), rd.get('diesel_issued',0)]
        for i, v in enumerate(row_vals, 1):
            ws3.cell(row=row, column=i, value=v)
        style_row(ws3, row, len(gen_hdr))
        row += 1

    set_col_widths(ws3, [11] + [14,10,12]*len(generators))

    # ---- Compressor Runtime ----
    ws4 = wb.create_sheet('Compressor Runtime')
    write_title(ws4, f'Compressor Runtime - {month_str}', site_name)

    compressors = c.execute(
        "SELECT id, name FROM equipment WHERE area='Utility' AND section='Compressed Air' ORDER BY name"
    ).fetchall()
    comp_hdr = ['Date']
    for comp in compressors:
        comp_hdr += [f"{comp['name']} - Hour Meter", f"{comp['name']} - Run Hrs"]

    row = 4
    for i, h in enumerate(comp_hdr, 1):
        ws4.cell(row=row, column=i, value=h)
    style_header(ws4, row, len(comp_hdr))
    row += 1

    comp_readings = c.execute("""
        SELECT ur.* FROM utility_readings ur
        JOIN equipment e ON e.id = ur.substation_id
        WHERE e.section='Compressed Air' AND ur.log_date BETWEEN ? AND ?
        ORDER BY ur.log_date, e.name
    """, (start, end)).fetchall()

    by_date_c = defaultdict(dict)
    for r in comp_readings:
        by_date_c[r['log_date']][r['equipment_id']] = dict(r)

    for log_date in sorted(by_date_c.keys()):
        day_data = by_date_c[log_date]
        row_vals = [log_date]
        for comp in compressors:
            rd = day_data.get(comp['id'], {})
            row_vals += [rd.get('hour_meter',''), rd.get('run_hours_today',0)]
        for i, v in enumerate(row_vals, 1):
            ws4.cell(row=row, column=i, value=v)
        style_row(ws4, row, len(comp_hdr))
        row += 1

    set_col_widths(ws4, [11] + [14,10]*len(compressors))

    # ---- PM Compliance ----
    ws5 = wb.create_sheet('PM Compliance')
    write_title(ws5, f'PM Compliance Archive - {month_str}', site_name)

    pm_hdr = ['Week Start', 'Artisan', 'Task', 'Equipment', 'Area',
              'Status', 'Completed At', 'Outcome', 'Condition', 'Parts Used']
    row = 4
    for i, h in enumerate(pm_hdr, 1):
        ws5.cell(row=row, column=i, value=h)
    style_header(ws5, row, len(pm_hdr))
    row += 1

    tasks = c.execute("""
        SELECT t.week_start, u.name, t.title, e.name as equip, e.area,
               t.status, t.completed_at, t.outcome, t.condition_found,
               GROUP_CONCAT(pp.part_desc || ' x' || pp.qty, ', ') as parts
        FROM   pm_tasks t
        JOIN   users u ON u.id = t.assigned_to
        LEFT JOIN equipment e ON e.id = t.equipment_id
        LEFT JOIN pm_parts pp ON pp.task_id = t.id
        WHERE  t.week_start BETWEEN ? AND ?
        GROUP  BY t.id ORDER BY t.week_start, u.name
    """, (start, end)).fetchall()

    for t in tasks:
        fill = GRN_FILL if t['status']=='done' else RED_FILL if t['status']=='missed' else None
        for i, v in enumerate(t, 1):
            ws5.cell(row=row, column=i, value=v or '')
        style_row(ws5, row, len(pm_hdr), fill)
        row += 1

    set_col_widths(ws5, [12,16,35,30,10,8,16,25,20,30])

    conn.close()
    fname = f"archive_{year}_{str(month).zfill(2)}.xlsx"
    fpath = os.path.join(REPORTS_DIR, fname)
    wb.save(fpath)
    return fpath


# ============================================================
# EMAIL HELPERS
# ============================================================
def get_setting_local(key, default=''):
    from database import get_setting
    return get_setting(key, default)

def send_email(subject, body, attachment_path, extra_recipients=None):
    MAIL_SENDER, MAIL_PASSWORD, MAIL_RECIPIENT = _get_mail_settings()
    if not MAIL_SENDER or not MAIL_PASSWORD or MAIL_PASSWORD == 'your_app_password_here':
        print("Email not configured - skipping send.")
        return False
    recipients = [MAIL_RECIPIENT] if MAIL_RECIPIENT else []
    if extra_recipients:
        for r in extra_recipients:
            if r and r not in recipients:
                recipients.append(r)
    if not recipients:
        print("No recipients configured.")
        return False
    try:
        for recipient in recipients:
            msg = MIMEMultipart()
            msg['From']    = MAIL_SENDER
            msg['To']      = recipient
            msg['Subject'] = subject
            msg.attach(MIMEText(body, 'plain', 'utf-8'))
            with open(attachment_path, 'rb') as f:
                part = MIMEBase('application', 'octet-stream')
                part.set_payload(f.read())
            encoders.encode_base64(part)
            part.add_header('Content-Disposition',
                            f'attachment; filename="{os.path.basename(attachment_path)}"')
            msg.attach(part)
            with smtplib.SMTP_SSL('smtp.gmail.com', 465) as server:
                server.login(MAIL_SENDER, MAIL_PASSWORD)
                server.sendmail(MAIL_SENDER, recipient, msg.as_string())
        logger.info(f"Email sent: {subject}")
        return True
    except Exception as e:
        logger.error(f"Email failed: {e}", exc_info=True)
        return False

def send_weekly_report():
    path       = generate_weekly_report()
    week_start = get_week_start()
    site_name  = get_setting_local('site_name', SITE_NAME)
    subject    = f"Weekly Maintenance Report - {week_start.strftime('%d %b %Y')}"
    body       = f"Please find attached the weekly maintenance report for {site_name}.\n\nWeek: {week_start.strftime('%d %b %Y')}"
    extra      = _get_recipients('weekly_report_recipients')
    return send_email(subject, body, path, extra), path

def send_monthly_archive(year=None, month=None):
    today = date.today()
    if not year or not month:
        first = today.replace(day=1)
        prev  = first - timedelta(days=1)
        year, month = prev.year, prev.month
    # Use the official template-aligned report when the template is available,
    # otherwise fall back to the internally-generated archive.
    try:
        from template_report import generate_template_report
        path, _stats = generate_template_report(year, month)
    except Exception as e:
        logger.warning(f"Template report unavailable ({e}); using internal archive.")
        path = generate_monthly_archive(year, month)
    month_s   = date(year, month, 1).strftime('%B %Y')
    site_name = get_setting_local('site_name', SITE_NAME)
    subject   = f"Monthly Equipment Availability & Reliability - {month_s}"
    body      = (f"Please find attached the Equipment Availability & Reliability "
                 f"report for {site_name}.\n\nMonth: {month_s}\n\n"
                 f"Note: open in Excel and allow pivot tables to refresh on load.")
    extra     = _get_recipients('monthly_report_recipients')
    return send_email(subject, body, path, extra), path

def purge_old_data(months=6):
    cutoff = (date.today() - timedelta(days=months*30)).isoformat()
    conn   = get_db()
    c      = conn.cursor()
    c.execute("DELETE FROM breakdowns WHERE DATE(logged_at) < ?", (cutoff,))
    c.execute("DELETE FROM pm_tasks WHERE week_start < ?", (cutoff,))
    c.execute("DELETE FROM activity_log WHERE DATE(logged_at) < ?", (cutoff,))
    conn.commit()
    conn.close()

def send_test_email():
    try:
        from database import get_setting
        sender    = get_setting('mail_sender')
        password  = get_setting('mail_password')
        recipient = get_setting('mail_recipient')
        if not sender or not password or not recipient:
            return False
        msg = MIMEText("Mill Maintenance Terminal - test email. Configuration is working correctly.", 'plain', 'utf-8')
        msg['Subject'] = 'Mill Terminal - Test Email'
        msg['From']    = sender
        msg['To']      = recipient
        with smtplib.SMTP_SSL('smtp.gmail.com', 465) as s:
            s.login(sender, password)
            s.sendmail(sender, recipient, msg.as_string())
        return True
    except Exception as e:
        logger.error(f"Test email failed: {e}", exc_info=True)
        return False
