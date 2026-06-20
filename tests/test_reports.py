"""
Unit tests for reports.py

Covers: style helpers, write_title, set_col_widths, get_setting_local,
        purge_old_data, _get_recipients.
"""

import os
import pytest
from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


# ============================================================
# Style helpers
# ============================================================

class TestStyleHeader:
    def test_applies_fill_and_font(self):
        from reports import style_header, HDR_FILL, HDR_FONT
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="Header")
        ws.cell(row=1, column=2, value="Header2")
        style_header(ws, 1, 2)
        assert ws.cell(row=1, column=1).fill == HDR_FILL
        assert ws.cell(row=1, column=1).font == HDR_FONT
        assert ws.cell(row=1, column=2).fill == HDR_FILL


class TestStyleRow:
    def test_applies_fill(self):
        from reports import style_row, RED_FILL, THIN
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="Data")
        style_row(ws, 1, 1, RED_FILL)
        assert ws.cell(row=1, column=1).fill == RED_FILL
        assert ws.cell(row=1, column=1).border == THIN

    def test_no_fill(self):
        from reports import style_row, THIN
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="Data")
        style_row(ws, 1, 1)
        assert ws.cell(row=1, column=1).border == THIN


class TestSetColWidths:
    def test_sets_widths(self):
        from reports import set_col_widths
        wb = Workbook()
        ws = wb.active
        set_col_widths(ws, [10, 20, 30])
        assert ws.column_dimensions["A"].width == 10
        assert ws.column_dimensions["B"].width == 20
        assert ws.column_dimensions["C"].width == 30


class TestWriteTitle:
    def test_sets_title(self):
        from reports import write_title
        wb = Workbook()
        ws = wb.active
        write_title(ws, "My Report", "Subtitle here")
        assert ws["A1"].value == "My Report"
        assert ws["A2"].value == "Subtitle here"
        assert ws.sheet_view.showGridLines is False

    def test_without_subtitle(self):
        from reports import write_title
        wb = Workbook()
        ws = wb.active
        write_title(ws, "My Report")
        assert ws["A1"].value == "My Report"
        assert ws["A2"].value is None


# ============================================================
# get_setting_local
# ============================================================

class TestGetSettingLocal:
    def test_returns_stored_value(self, db_conn):
        from database import set_setting
        from reports import get_setting_local
        set_setting("site_name", "Test Mill")
        assert get_setting_local("site_name") == "Test Mill"

    def test_returns_default(self, db_conn):
        from reports import get_setting_local
        assert get_setting_local("nonexistent", "default_val") == "default_val"


# ============================================================
# _get_recipients
# ============================================================

class TestGetRecipients:
    def test_returns_empty_list_default(self, db_conn):
        from reports import _get_recipients
        result = _get_recipients("weekly_report_recipients")
        assert result == []

    def test_returns_parsed_json(self, db_conn):
        from database import set_setting
        from reports import _get_recipients
        import json
        set_setting("weekly_report_recipients", json.dumps(["a@test.com", "b@test.com"]))
        result = _get_recipients("weekly_report_recipients")
        assert result == ["a@test.com", "b@test.com"]

    def test_handles_invalid_json(self, db_conn):
        from database import set_setting
        from reports import _get_recipients
        set_setting("weekly_report_recipients", "not-json")
        result = _get_recipients("weekly_report_recipients")
        assert result == []


# ============================================================
# purge_old_data
# ============================================================

class TestPurgeOldData:
    def test_purges_old_breakdowns(self, db_conn):
        from reports import purge_old_data
        conn = db_conn
        conn.execute("INSERT INTO users (id, name, role) VALUES (1, 'Test', 'artisan')")
        conn.execute("INSERT INTO equipment (id, name, area) VALUES (1, 'M1', 'Wetmill')")

        old_date = (date.today() - timedelta(days=365)).isoformat()
        recent_date = date.today().isoformat()

        conn.execute(
            "INSERT INTO breakdowns (equipment_id, supervisor_id, logged_at, status) "
            "VALUES (1, 1, ?, 'completed')", (old_date,)
        )
        conn.execute(
            "INSERT INTO breakdowns (equipment_id, supervisor_id, logged_at, status) "
            "VALUES (1, 1, ?, 'completed')", (recent_date,)
        )
        conn.commit()

        purge_old_data(months=6)

        count = conn.execute("SELECT COUNT(*) FROM breakdowns").fetchone()[0]
        assert count == 1

    def test_purges_old_pm_tasks(self, db_conn):
        from reports import purge_old_data
        conn = db_conn
        conn.execute("INSERT INTO users (id, name, role) VALUES (1, 'Test', 'artisan')")
        conn.execute("INSERT INTO equipment (id, name, area) VALUES (1, 'M1', 'Wetmill')")

        old_week = (date.today() - timedelta(days=365)).isoformat()
        recent_week = date.today().isoformat()

        conn.execute(
            "INSERT INTO pm_tasks (title, equipment_id, assigned_to, week_start, status) "
            "VALUES ('Old', 1, 1, ?, 'done')", (old_week,)
        )
        conn.execute(
            "INSERT INTO pm_tasks (title, equipment_id, assigned_to, week_start, status) "
            "VALUES ('Recent', 1, 1, ?, 'pending')", (recent_week,)
        )
        conn.commit()

        purge_old_data(months=6)

        tasks = conn.execute("SELECT title FROM pm_tasks").fetchall()
        assert len(tasks) == 1
        assert tasks[0]["title"] == "Recent"

    def test_purges_old_activity_log(self, db_conn):
        from reports import purge_old_data
        conn = db_conn

        old_date = (date.today() - timedelta(days=365)).isoformat()
        recent_date = date.today().isoformat()

        conn.execute(
            "INSERT INTO activity_log (user_name, action_type, description, logged_at) "
            "VALUES ('Test', 'login', 'Old login', ?)", (old_date,)
        )
        conn.execute(
            "INSERT INTO activity_log (user_name, action_type, description, logged_at) "
            "VALUES ('Test', 'login', 'Recent login', ?)", (recent_date,)
        )
        conn.commit()

        purge_old_data(months=6)

        count = conn.execute("SELECT COUNT(*) FROM activity_log").fetchone()[0]
        assert count == 1


# ============================================================
# send_email (no actual email sent)
# ============================================================

class TestSendEmail:
    def test_returns_false_without_config(self, db_conn):
        from reports import send_email
        result = send_email("Test", "Body", "/tmp/fake.xlsx")
        assert result is False

    def test_returns_false_with_placeholder_password(self, db_conn):
        from database import set_setting
        from reports import send_email
        set_setting("mail_sender", "test@test.com")
        set_setting("mail_password", "your_app_password_here")
        set_setting("mail_recipient", "recipient@test.com")
        result = send_email("Test", "Body", "/tmp/fake.xlsx")
        assert result is False
