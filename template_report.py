"""
template_report.py - V10
Populates the official Equipment Availability & Reliability template
by injecting data from the Mill Maintenance Terminal database into
the exact input cells, leaving all formulas and dashboards intact.

Approach:
- Copy blank template (never modify the master)
- Inject breakdown rows into Daily Downtime log
- Inject boiler readings into Boiler operations
- Inject generator/compressor readings into runtime sheets
- Return populated file — formulas calculate everything else
"""

import os
import shutil
from datetime import date, timedelta, datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))

TEMPLATE_CANDIDATES = [
    os.path.join(HERE, 'templates', 'Equipment_availability_template.xlsx'),
    os.path.join(HERE, 'templates', 'Equipment_availability___reliability_-_Template__Monthly_.xlsx'),
]

REPORTS_DIR = os.path.join(HERE, 'reports')
os.makedirs(REPORTS_DIR, exist_ok=True)


def _find_template():
    for path in TEMPLATE_CANDIDATES:
        if os.path.exists(path):
            return path
    raise FileNotFoundError(
        "Template not found. Place the blank template at:\n"
        f"  {TEMPLATE_CANDIDATES[0]}"
    )


def _build_equipment_map(conn):
    """Build mapping from our equipment names/IDs to template compound IDs."""
    rows = conn.execute(
        "SELECT id, name, compound_id, area, section FROM equipment WHERE active=1"
    ).fetchall()
    by_id = {}
    by_name = {}
    for r in rows:
        by_id[r['id']] = dict(r)
        by_name[r['name'].lower()] = dict(r)
    return by_id, by_name


def _get_compound_id(equip_id, equip_name, by_id, by_name):
    """Return compound ID for template, e.g. 'BDS-002 - BE-007'."""
    if equip_id and equip_id in by_id:
        cid = by_id[equip_id].get('compound_id')
        if cid:
            return cid
    if equip_name:
        match = by_name.get(equip_name.lower())
        if match and match.get('compound_id'):
            return match['compound_id']
    return equip_name or ''


def generate_template_report(year, month):
    """
    Populate the official template for the given month/year.
    Returns (file_path, stats_dict).
    """
    from database import get_db

    template_path = _find_template()
    fname = f"availability_{year}_{str(month).zfill(2)}.xlsx"
    fpath = os.path.join(REPORTS_DIR, fname)
    shutil.copy2(template_path, fpath)

    wb = load_workbook(fpath)
    conn = get_db()
    by_id, by_name = _build_equipment_map(conn)

    start = date(year, month, 1).isoformat()
    # Last day of month
    if month == 12:
        end = date(year + 1, 1, 1) - timedelta(days=1)
    else:
        end = date(year, month + 1, 1) - timedelta(days=1)
    end = end.isoformat()

    stats = {'breakdowns': 0, 'boiler_days': 0, 'gen_days': 0, 'comp_days': 0}

    # ----------------------------------------------------------------
    # 1. Daily Downtime log
    # Input cols: A=Date, C=Equipment ID, E=Description,
    #             F=Start, G=End, I=Failure Type,
    #             K=Failed Component, L=Failure mode, Q=Comments
    # Rows start at 2 (row 1 is header)
    # ----------------------------------------------------------------
    ws_dd = wb['Daily Downtime log']

    breakdowns = conn.execute("""
        SELECT b.*, e.name as equip_name, e.compound_id,
               u.name as artisan_name
        FROM breakdowns b
        LEFT JOIN equipment e ON e.id = b.equipment_id
        LEFT JOIN users u ON u.id = b.artisan_id
        WHERE b.log_date BETWEEN ? AND ?
        ORDER BY b.log_date, b.sup_time_start
    """, (start, end)).fetchall()

    row = 2
    for b in breakdowns:
        compound = b['compound_id'] or _get_compound_id(
            b['equipment_id'], b['equip_name'], by_id, by_name)

        # Format times as fractions of day (Excel time format)
        def time_to_excel(t):
            if not t: return None
            try:
                parts = str(t).split(':')
                h, m = int(parts[0]), int(parts[1])
                return (h * 60 + m) / 1440
            except:
                return None

        # Template columns:
        # A=Date, B=Dept, C=Equipment ID, D=Equipment Name, E=Description,
        # F=Start, G=End, I=Failure Type, K=Component, L=Failure Mode, Q=Comments
        
        ws_dd.cell(row=row, column=1).value  = b['log_date']
        ws_dd.cell(row=row, column=2).value  = b.get('area') or 'Wetmill'
        ws_dd.cell(row=row, column=3).value  = compound
        ws_dd.cell(row=row, column=4).value  = b['equip_name'] or compound.split(' - ')[-1]
        ws_dd.cell(row=row, column=5).value  = (b['art_description'] or
                                                 b['sup_notes'] or '')
        ws_dd.cell(row=row, column=6).value  = time_to_excel(b['sup_time_start'])
        ws_dd.cell(row=row, column=6).number_format = 'HH:MM'
        ws_dd.cell(row=row, column=7).value  = time_to_excel(
            b['art_time_end'] or b['sup_time_end'])
        ws_dd.cell(row=row, column=7).number_format = 'HH:MM'
        ws_dd.cell(row=row, column=9).value  = (b['final_failure_type'] or
                                                 b['sup_failure_type'] or '')
        ws_dd.cell(row=row, column=11).value = (b['art_failed_component'] or
                                                 b['sup_component'] or '')
        ws_dd.cell(row=row, column=12).value = (b['art_failure_mode'] or
                                                 b['sup_failure_mode'] or '')
        ws_dd.cell(row=row, column=17).value = (b['art_repair_action'] or '')
        row += 1
        stats['breakdowns'] += 1

    # ----------------------------------------------------------------
    # 2. Boiler operations
    # Row 5 = headers, Row 6 = baseline (prev month), Row 7+ = daily data
    # Input: A=Date, B=B1 Morning TDS, C=B1 Night TDS, E=B1 Blowdowns
    #        F=B2 Morning, G=B2 Night, I=B2 Blowdowns
    #        J=B3 Morning, K=B3 Night, M=B3 Blowdowns
    #        P=Condensate meter, R=Makeup meter
    #        X=Softener Day, Y=Softener Night, Z=Salt bags
    #        AA=Lost time, AB=Incident, AC=Comments
    # ----------------------------------------------------------------
    if 'Boiler operations' in wb.sheetnames:
        ws_bo = wb['Boiler operations']

        boiler_readings = conn.execute("""
            SELECT br.*, e.name as boiler_name
            FROM boiler_readings br
            LEFT JOIN equipment e ON e.id = br.boiler_id
            WHERE br.log_date BETWEEN ? AND ?
            ORDER BY br.log_date, e.name
        """, (start, end)).fetchall()

        # Group by date
        by_date = {}
        for r in boiler_readings:
            d = r['log_date']
            if d not in by_date:
                by_date[d] = []
            by_date[d].append(dict(r))

        row = 7  # data starts at row 7
        for log_date in sorted(by_date.keys()):
            readings = by_date[log_date]
            ws_bo.cell(row=row, column=1).value = log_date

            for i, r in enumerate(readings[:3]):
                base = i * 4  # B1=cols 2-5, B2=cols 6-9, B3=cols 10-13
                if i == 0:
                    ws_bo.cell(row=row, column=2).value = r.get('tds_morning')
                    ws_bo.cell(row=row, column=3).value = r.get('tds_night')
                    ws_bo.cell(row=row, column=5).value = r.get('blowdowns_total')
                elif i == 1:
                    ws_bo.cell(row=row, column=6).value = r.get('tds_morning')
                    ws_bo.cell(row=row, column=7).value = r.get('tds_night')
                    ws_bo.cell(row=row, column=9).value = r.get('blowdowns_total')
                elif i == 2:
                    ws_bo.cell(row=row, column=10).value = r.get('tds_morning')
                    ws_bo.cell(row=row, column=11).value = r.get('tds_night')
                    ws_bo.cell(row=row, column=13).value = r.get('blowdowns_total')

            # Use first reading for shared columns
            if readings:
                r0 = readings[0]
                ws_bo.cell(row=row, column=16).value = r0.get('condensate_meter_reading')
                ws_bo.cell(row=row, column=18).value = r0.get('makeup_meter_reading')
                ws_bo.cell(row=row, column=24).value = r0.get('softener_day')
                ws_bo.cell(row=row, column=25).value = r0.get('softener_night')
                ws_bo.cell(row=row, column=26).value = r0.get('salt_bags')
                ws_bo.cell(row=row, column=27).value = r0.get('lost_time_mins')
                ws_bo.cell(row=row, column=28).value = r0.get('downtime_incident')
                ws_bo.cell(row=row, column=29).value = r0.get('notes')

            row += 1
            stats['boiler_days'] += 1

    # ----------------------------------------------------------------
    # 3. Generator runtime
    # 4 generators, baseline row 4, data rows 5+
    # WM: A=Date, B=Hour meter | DM: G=Date, H=Hour meter
    # BH: J=Date, K=Hour meter | Admin: P=Date, Q=Hour meter
    # ----------------------------------------------------------------
    if 'Generator runtime' in wb.sheetnames:
        ws_gr = wb['Generator runtime']

        gen_readings = conn.execute("""
            SELECT ur.*, e.name as equip_name, e.compound_id
            FROM utility_readings ur
            JOIN equipment e ON e.id = ur.substation_id
            WHERE (e.name LIKE '%Generator%' OR e.name LIKE '%Gen%')
            AND ur.log_date BETWEEN ? AND ?
            ORDER BY ur.log_date, e.name
        """, (start, end)).fetchall()

        # Map generator to column
        GEN_COLS = {
            'wetmill': (1, 2),   # Date col A, Hour meter col B
            'drymill': (7, 8),   # Date col G, Hour meter col H
            'boiler':  (10, 11), # Date col J, Hour meter col K
            'admin':   (16, 17), # Date col P, Hour meter col Q
        }

        by_date_gen = {}
        for r in gen_readings:
            d = r['log_date']
            if d not in by_date_gen:
                by_date_gen[d] = {}
            name = (r['equip_name'] or '').lower()
            if 'wetmill' in name or '523' in name or '2mva' in name or '2000' in name:
                by_date_gen[d]['wetmill'] = dict(r)
            elif 'drymill' in name or '18' in name or '1mva' in name or '1000' in name:
                by_date_gen[d]['drymill'] = dict(r)
            elif 'boiler' in name or '148' in name or '800' in name:
                by_date_gen[d]['boiler'] = dict(r)
            elif 'admin' in name or '40' in name or 'goscor' in name:
                by_date_gen[d]['admin'] = dict(r)

        row = 5
        for log_date in sorted(by_date_gen.keys()):
            day = by_date_gen[log_date]
            for key, (dcol, hcol) in GEN_COLS.items():
                if key in day:
                    ws_gr.cell(row=row, column=dcol).value = log_date
                    ws_gr.cell(row=row, column=hcol).value = day[key].get('hour_meter')
            row += 1
            stats['gen_days'] += 1

    # ----------------------------------------------------------------
    # 4. Compressor runtime
    # 4 compressors, baseline row 4, data rows 5+
    # KX-01: A=Date, B=Hour | KX-02: E=Date, F=Hour
    # KX-03: I=Date, J=Hour | KX-04: M=Date, N=Hour
    # ----------------------------------------------------------------
    if 'Compressor runtime' in wb.sheetnames:
        ws_cr = wb['Compressor runtime']

        comp_readings = conn.execute("""
            SELECT ur.*, e.name as equip_name
            FROM utility_readings ur
            JOIN equipment e ON e.id = ur.substation_id
            WHERE (e.name LIKE '%Compressor%' OR e.name LIKE '%KX%')
            AND ur.log_date BETWEEN ? AND ?
            ORDER BY ur.log_date, e.name
        """, (start, end)).fetchall()

        COMP_COLS = [
            (1, 2),   # KX-01: col A, B
            (5, 6),   # KX-02: col E, F
            (9, 10),  # KX-03: col I, J
            (13, 14), # KX-04: col M, N
        ]

        by_date_comp = {}
        for r in comp_readings:
            d = r['log_date']
            if d not in by_date_comp:
                by_date_comp[d] = []
            by_date_comp[d].append(dict(r))

        row = 5
        for log_date in sorted(by_date_comp.keys()):
            comps = by_date_comp[log_date][:4]
            for i, r in enumerate(comps):
                if i < len(COMP_COLS):
                    dcol, hcol = COMP_COLS[i]
                    ws_cr.cell(row=row, column=dcol).value = log_date
                    ws_cr.cell(row=row, column=hcol).value = r.get('hour_meter')
            row += 1
            stats['comp_days'] += 1

    conn.close()
    wb.save(fpath)

    return fpath, stats