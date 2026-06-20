"""
Mill Maintenance Terminal - Improved Version
With error handling, structured logging, and better database management
"""

from flask import Flask, request, jsonify, send_from_directory, send_file, session
from werkzeug.security import generate_password_hash, check_password_hash
import json, os, secrets, threading, time, logging
from datetime import datetime, date, timedelta
from database import get_db_context, init_db, seed_data, get_setting, set_setting, DB_FILE
from scheduler import generate_weekly_tasks, get_tasks_for_artisan, mark_task_done, get_week_start
from reports import generate_weekly_report, send_weekly_report, send_monthly_archive

# ---- Logging Setup ----
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('mill.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

app = Flask(__name__, static_folder=os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static'))
app.secret_key = os.environ.get('FLASK_SECRET_KEY') or secrets.token_hex(32)

# ---- Global Error Handlers ----
@app.errorhandler(400)
def bad_request(error):
    """Handle 400 Bad Request errors."""
    logger.warning(f"Bad request: {error.description}")
    return jsonify({'status': 'error', 'message': 'Invalid request: ' + str(error.description)}), 400

@app.errorhandler(404)
def not_found(error):
    """Handle 404 Not Found errors."""
    logger.warning(f"Not found: {request.path}")
    return jsonify({'status': 'error', 'message': 'Endpoint not found'}), 404

@app.errorhandler(500)
def server_error(error):
    """Handle 500 Internal Server errors."""
    logger.error(f"Server error: {error}", exc_info=True)
    return jsonify({'status': 'error', 'message': 'Internal server error. Please try again later.'}), 500

@app.errorhandler(Exception)
def handle_exception(error):
    """Catch all unhandled exceptions."""
    logger.error(f"Unhandled exception: {error}", exc_info=True)
    return jsonify({'status': 'error', 'message': 'An unexpected error occurred'}), 500

# ---- Initialization ----
init_db()
seed_data()

try:
    generate_weekly_tasks()
    logger.info("Weekly task generation completed on startup")
except Exception as e:
    logger.error(f"Weekly task generation on startup failed: {e}", exc_info=True)

# ---- Helper Functions ----
def get_days_arg(default=7, maximum=366):
    """Parse and validate 'days' query parameter."""
    try:
        days = int(request.args.get('days', default))
        return max(1, min(days, maximum))
    except (TypeError, ValueError):
        logger.debug(f"Invalid days parameter, using default: {default}")
        return default

def log_action(user, action, desc):
    """Log an activity to the database."""
    try:
        with get_db_context() as conn:
            conn.execute(
                "INSERT INTO activity_log (user_name,action_type,description) VALUES (?,?,?)",
                (user, action, desc)
            )
            conn.commit()
            logger.info(f"Action logged: {action} by {user}")
    except Exception as e:
        logger.error(f"Failed to log action: {e}", exc_info=True)

def safe_json(request_obj):
    """Safely get JSON from request."""
    try:
        return request_obj.get_json(silent=True) or {}
    except Exception as e:
        logger.debug(f"Failed to parse JSON: {e}")
        return {}

# ---- Routes: Static Files ----
@app.route('/')
def index():
    """Serve the main HTML page."""
    return send_from_directory(app.static_folder, 'index.html')

# ---- Routes: Users & PIN ----
@app.route('/api/users')
def get_users():
    """Get all active users."""
    try:
        with get_db_context() as conn:
            users = conn.execute(
                "SELECT id,name,role,trade,CASE WHEN pin IS NOT NULL THEN 1 ELSE 0 END as has_pin "
                "FROM users WHERE active=1 ORDER BY role,name"
            ).fetchall()
            return jsonify([dict(u) for u in users])
    except Exception as e:
        logger.error(f"Error fetching users: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': 'Failed to fetch users'}), 500

@app.route('/api/users/verify-pin', methods=['POST'])
def verify_pin():
    """Verify user PIN."""
    d = safe_json(request)
    
    if 'user_id' not in d or 'pin' not in d:
        return jsonify({'valid': False, 'error': 'user_id and pin are required'}), 400
    
    try:
        with get_db_context() as conn:
            user = conn.execute(
                "SELECT id,name,role,pin FROM users WHERE id=?",
                (d['user_id'],)
            ).fetchone()
        
        if not user:
            logger.warning(f"Failed PIN verification for user {d['user_id']}")
            return jsonify({'valid': False})

        stored_pin = user['pin']
        if stored_pin is None:
            valid = True
        elif stored_pin.startswith(('pbkdf2:', 'scrypt:')):
            valid = check_password_hash(stored_pin, d['pin'])
        else:
            valid = (stored_pin == d['pin'])
            if valid:
                with get_db_context() as conn:
                    conn.execute("UPDATE users SET pin=? WHERE id=?",
                                 (generate_password_hash(d['pin']), user['id']))
                    conn.commit()

        if valid:
            return jsonify({'valid': True, 'user': {'id': user['id'], 'name': user['name'], 'role': user['role']}})
        else:
            logger.warning(f"Failed PIN verification for user {d['user_id']}")
            return jsonify({'valid': False})
    except Exception as e:
        logger.error(f"Error verifying PIN: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': 'PIN verification failed'}), 500

@app.route('/api/users/<int:uid>/pin', methods=['POST'])
def set_pin(uid):
    """Set or update user PIN."""
    d = safe_json(request)
    pin = str(d.get('pin', '')).strip()
    
    if pin and (len(pin) != 4 or not pin.isdigit()):
        return jsonify({'status': 'error', 'message': 'PIN must contain exactly 4 digits'}), 400
    
    try:
        hashed_pin = generate_password_hash(pin) if pin else None
        with get_db_context() as conn:
            conn.execute("UPDATE users SET pin=? WHERE id=?", (hashed_pin, uid))
            conn.commit()
        
        log_action('Admin', 'pin_updated', f"PIN updated for user {uid}")
        return jsonify({'status': 'ok'})
    except Exception as e:
        logger.error(f"Error setting PIN: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': 'Failed to set PIN'}), 500

@app.route('/api/users', methods=['POST'])
def add_user():
    """Add a new user."""
    d = safe_json(request)
    name = str(d.get('name', '')).strip()
    role = d.get('role')
    
    valid_roles = {'manager', 'admin', 'supervisor', 'artisan', 'sawshop', 'team_leader', 'picker', 'storeman', 'stores_admin'}
    if not name or role not in valid_roles:
        return jsonify({'status': 'error', 'message': f'Valid name and role required. Valid roles: {valid_roles}'}), 400

    creator_role = d.get('creator_role', '')
    if creator_role == 'stores_admin' and role not in ('picker', 'storeman'):
        return jsonify({'status': 'error', 'message': 'Stores admin can only add picker or storeman roles'}), 403

    try:
        raw_pin = d.get('pin')
        hashed_pin = generate_password_hash(str(raw_pin)) if raw_pin else None
        with get_db_context() as conn:
            conn.execute(
                "INSERT INTO users (name,role,pin) VALUES (?,?,?)",
                (name, role, hashed_pin)
            )
            conn.commit()
        
        log_action('Admin', 'user_added', f"User added: {name} ({role})")
        return jsonify({'status': 'ok'})
    except Exception as e:
        logger.error(f"Error adding user: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': 'Failed to add user'}), 500

@app.route('/api/users/<int:uid>', methods=['DELETE'])
def deactivate_user(uid):
    """Deactivate a user."""
    try:
        with get_db_context() as conn:
            conn.execute("UPDATE users SET active=0 WHERE id=?", (uid,))
            conn.commit()
        
        log_action('Admin', 'user_deactivated', f"User {uid} deactivated")
        return jsonify({'status': 'ok'})
    except Exception as e:
        logger.error(f"Error deactivating user: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': 'Failed to deactivate user'}), 500

# ---- Routes: Equipment ----
@app.route('/api/equipment')
def get_equipment():
    """Get all active equipment."""
    try:
        with get_db_context() as conn:
            eq = conn.execute(
                "SELECT * FROM equipment WHERE active=1 ORDER BY area,section,name"
            ).fetchall()
            return jsonify([dict(e) for e in eq])
    except Exception as e:
        logger.error(f"Error fetching equipment: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': 'Failed to fetch equipment'}), 500

@app.route('/api/equipment/hierarchy')
def get_equipment_hierarchy():
    """Get equipment organized by area and section."""
    try:
        with get_db_context() as conn:
            eq = conn.execute(
                "SELECT id,name,area,section,equipment_id FROM equipment WHERE active=1 ORDER BY area,section,name"
            ).fetchall()
        
        hierarchy = {}
        for e in eq:
            area = e['area']
            section = e['section'] or 'General'
            if area not in hierarchy:
                hierarchy[area] = {}
            if section not in hierarchy[area]:
                hierarchy[area][section] = []
            hierarchy[area][section].append({
                'id': e['id'],
                'name': e['name'],
                'equipment_id': e['equipment_id']
            })
        
        return jsonify(hierarchy)
    except Exception as e:
        logger.error(f"Error fetching equipment hierarchy: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': 'Failed to fetch hierarchy'}), 500

# ---- Routes: Failure Modes ----
@app.route('/api/failure-modes')
def get_failure_modes():
    """Get all active failure modes."""
    try:
        with get_db_context() as conn:
            modes = conn.execute(
                "SELECT id, mode_name, category FROM failure_modes WHERE active=1 ORDER BY category, mode_name"
            ).fetchall()
            return jsonify([dict(m) for m in modes])
    except Exception as e:
        logger.error(f"Error fetching failure modes: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': 'Failed to fetch failure modes'}), 500

@app.route('/api/failure-modes', methods=['POST'])
def add_failure_mode():
    """Add a new failure mode."""
    d = safe_json(request)
    
    if not d.get('mode_name'):
        return jsonify({'status': 'error', 'message': 'mode_name is required'}), 400
    
    try:
        with get_db_context() as conn:
            conn.execute(
                "INSERT INTO failure_modes (mode_name, category) VALUES (?, ?)",
                (d['mode_name'], d.get('category', 'Mechanical'))
            )
            conn.commit()
        
        log_action('Admin', 'failure_mode_added', f"Failure mode added: {d['mode_name']}")
        return jsonify({'status': 'ok'})
    except Exception as e:
        logger.error(f"Error adding failure mode: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': 'Failed to add failure mode'}), 500

# ---- Routes: Tasks ----
@app.route('/api/tasks/unallocated')
def get_unallocated_tasks():
    """Get PM tasks not yet assigned."""
    try:
        with get_db_context() as conn:
            tasks = conn.execute(
                "SELECT id, title, description, equipment_id, status, created_at "
                "FROM pm_tasks WHERE (assigned_to IS NULL OR assigned_to=0) AND status='pending' "
                "ORDER BY created_at DESC"
            ).fetchall()
            return jsonify([dict(t) for t in tasks])
    except Exception as e:
        logger.error(f"Error fetching unallocated tasks: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': 'Failed to fetch tasks'}), 500

# ---- Routes: Settings ----
@app.route('/api/settings')
def get_settings():
    """Get all application settings."""
    try:
        with get_db_context() as conn:
            rows = conn.execute("SELECT key,value FROM settings").fetchall()
        
        result = {r['key']: r['value'] for r in rows}
        # Mask password in response
        if result.get('mail_password'):
            result['mail_password_set'] = True
            result['mail_password'] = '\u2022' * 16
        
        return jsonify(result)
    except Exception as e:
        logger.error(f"Error fetching settings: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': 'Failed to fetch settings'}), 500

@app.route('/api/settings', methods=['POST'])
def save_settings():
    """Save application settings."""
    d = safe_json(request)
    
    try:
        for key, val in d.items():
            # Don't overwrite password if masked
            if key == 'mail_password' and str(val).startswith('\u2022'):
                continue
            set_setting(key, str(val))
        
        log_action('Admin', 'settings_updated', 'Settings updated')
        return jsonify({'status': 'ok'})
    except Exception as e:
        logger.error(f"Error saving settings: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': 'Failed to save settings'}), 500

@app.route('/api/settings/test-email', methods=['POST'])
def test_email():
    """Send a test email to verify configuration."""
    try:
        from reports import send_test_email
        ok = send_test_email()
        status = 'sent' if ok else 'failed'
        logger.info(f"Test email result: {status}")
        return jsonify({'status': status})
    except Exception as e:
        logger.error(f"Test email failed: {e}", exc_info=True)
        return jsonify({'status': 'failed', 'error': 'Email configuration error'}), 500

# ---- Routes: Shift Confirmation ----
@app.route('/api/shift/check/<int:user_id>')
def check_shift(user_id):
    """Check if user has confirmed shift for current week."""
    try:
        week_start = get_week_start().isoformat()
        with get_db_context() as conn:
            row = conn.execute(
                "SELECT shift FROM shift_confirmations WHERE user_id=? AND week_start=?",
                (user_id, week_start)
            ).fetchone()
        
        return jsonify({
            'confirmed': row is not None,
            'shift': row['shift'] if row else None
        })
    except Exception as e:
        logger.error(f"Error checking shift: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': 'Failed to check shift'}), 500

@app.route('/api/shift/confirm', methods=['POST'])
def confirm_shift():
    """Confirm user shift for the week."""
    d = safe_json(request)
    
    if 'user_id' not in d or 'shift' not in d:
        return jsonify({'status': 'error', 'message': 'user_id and shift are required'}), 400
    
    try:
        week_start = get_week_start().isoformat()
        with get_db_context() as conn:
            conn.execute(
                "INSERT OR REPLACE INTO shift_confirmations (user_id,week_start,shift) VALUES (?,?,?)",
                (d['user_id'], week_start, d['shift'])
            )
            conn.commit()
        
        logger.info(f"Shift confirmed: user {d['user_id']}, shift {d['shift']}")
        return jsonify({'status': 'ok'})
    except Exception as e:
        logger.error(f"Error confirming shift: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': 'Failed to confirm shift'}), 500

# ---- Routes: Reports ----
@app.route('/api/report/weekly/download')
def download_weekly():
    """Download weekly report."""
    try:
        path = generate_weekly_report()
        logger.info(f"Weekly report downloaded: {path}")
        return send_file(path, as_attachment=True)
    except Exception as e:
        logger.error(f"Error generating weekly report: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': 'Failed to generate report'}), 500

@app.route('/api/report/weekly/email', methods=['POST'])
def email_weekly():
    """Email weekly report."""
    try:
        ok, path = send_weekly_report()
        status = 'sent' if ok else 'saved'
        logger.info(f"Weekly report emailed: {status}")
        return jsonify({'status': status, 'file': os.path.basename(path)})
    except Exception as e:
        logger.error(f"Error emailing weekly report: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': 'Failed to email report'}), 500

@app.route('/api/report/archive/download')
def download_archive():
    """Download monthly archive report."""
    try:
        year = int(request.args.get('year', date.today().year))
        month = int(request.args.get('month', date.today().month))
        path = send_monthly_archive(year, month)[1]
        logger.info(f"Archive report downloaded: {year}-{month:02d}")
        return send_file(path, as_attachment=True)
    except Exception as e:
        logger.error(f"Error downloading archive: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': 'Failed to download archive'}), 500

@app.route('/api/report/archive/email', methods=['POST'])
def email_archive():
    """Email monthly archive report."""
    d = safe_json(request)
    try:
        year = int(d.get('year', date.today().year))
        month = int(d.get('month', date.today().month))
        ok, path = send_monthly_archive(year, month)
        status = 'sent' if ok else 'saved'
        logger.info(f"Archive emailed: {year}-{month:02d} ({status})")
        return jsonify({'status': status, 'file': os.path.basename(path)})
    except Exception as e:
        logger.error(f"Error emailing archive: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': 'Failed to email archive'}), 500

# ============================================================
# v10.1 ADDED ROUTES — Jobs, Breakdowns, Requisitions, Stores
# ============================================================

# ---- Helper: week start ----
def _week_start_iso():
    return get_week_start().isoformat()

# ============================================================
# TASKS / JOBS  (Admin allocates → Artisan does → back to Admin)
# ============================================================
@app.route('/api/tasks/allocate', methods=['POST'])
def allocate_task():
    """Admin/manager allocates a job to an artisan. Appears on artisan board immediately."""
    d = safe_json(request)
    title = str(d.get('title', '')).strip()
    assigned_to = d.get('assigned_to')
    if not title or not assigned_to:
        return jsonify({'status': 'error', 'message': 'title and assigned_to are required'}), 400
    try:
        week_start = _week_start_iso()
        with get_db_context() as conn:
            cur = conn.execute("""
                INSERT INTO pm_tasks
                (title, equipment_id, assigned_to, week_start, due_date, status,
                 task_type, description, recurring, frequency, created_at)
                VALUES (?,?,?,?,?,'pending','pm',?,?,?,?)
            """, (
                title, d.get('equipment_id'), int(assigned_to), week_start,
                (get_week_start() + timedelta(days=6)).isoformat(),
                d.get('description', ''), int(d.get('recurring', 0) or 0),
                d.get('frequency'), datetime.now().isoformat()
            ))
            task_id = cur.lastrowid
            # Optional checklist items
            for i, item in enumerate(d.get('checklist_items', []) or []):
                if item:
                    conn.execute(
                        "INSERT INTO pm_checklist_items (task_id,item,sort_order) VALUES (?,?,?)",
                        (task_id, item, i)
                    )
            conn.commit()
        log_action(d.get('allocated_by_name', 'Admin'), 'task_allocated', f"Task '{title}' -> user {assigned_to}")
        return jsonify({'status': 'ok', 'id': task_id})
    except Exception as e:
        logger.error(f"Error allocating task: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': 'Failed to allocate task'}), 500


@app.route('/api/tasks/<int:user_id>')
def get_user_tasks(user_id):
    """Get an artisan's tasks for the week + breakdowns awaiting their input."""
    try:
        week_start = _week_start_iso()
        with get_db_context() as conn:
            tasks = conn.execute("""
                SELECT t.*, e.name as equip_name, e.area, e.section
                FROM pm_tasks t
                LEFT JOIN equipment e ON e.id = t.equipment_id
                WHERE t.assigned_to=? AND t.week_start=?
                  AND t.status != 'done' AND COALESCE(t.suspended,0)=0
                ORDER BY t.created_at DESC
            """, (user_id, week_start)).fetchall()
            result = []
            for t in tasks:
                td = dict(t)
                items = conn.execute(
                    "SELECT * FROM pm_checklist_items WHERE task_id=? ORDER BY sort_order",
                    (t['id'],)
                ).fetchall()
                td['checklist'] = [dict(i) for i in items]
                result.append(td)

            # Breakdowns awaiting this artisan
            bds = conn.execute("""
                SELECT b.*, e.name as equip_name, e.area,
                       s.name as supervisor_name
                FROM breakdowns b
                LEFT JOIN equipment e ON e.id = b.equipment_id
                LEFT JOIN users s ON s.id = b.supervisor_id
                WHERE b.artisan_id=? AND b.status='supervisor_logged'
                ORDER BY b.logged_at DESC
            """, (user_id,)).fetchall()
            pending_bds = [dict(b) for b in bds]

        return jsonify({'tasks': result, 'pending_breakdowns': pending_bds})
    except Exception as e:
        logger.error(f"Error fetching tasks for user {user_id}: {e}", exc_info=True)
        return jsonify({'tasks': [], 'pending_breakdowns': []}), 500


@app.route('/api/tasks/<int:task_id>/complete', methods=['POST'])
def complete_task(task_id):
    """Artisan marks a PM task done. Moves to DidIt + admin dashboard."""
    d = safe_json(request)
    try:
        with get_db_context() as conn:
            conn.execute("""
                UPDATE pm_tasks
                SET status='done', completed_by=?, completed_at=?,
                    outcome=?, condition_found=?, notes=?
                WHERE id=?
            """, (
                d.get('completed_by'), datetime.now().isoformat(),
                d.get('outcome', ''), d.get('condition', ''),
                d.get('notes', ''), task_id
            ))
            # checklist states
            for item_id, checked in (d.get('checklist', {}) or {}).items():
                try:
                    conn.execute(
                        "UPDATE pm_checklist_items SET checked=? WHERE id=?",
                        (1 if checked else 0, int(item_id))
                    )
                except (ValueError, TypeError):
                    pass
            # parts used -> reorder
            for p in (d.get('parts', []) or []):
                if p.get('desc'):
                    conn.execute(
                        "INSERT INTO pm_parts (task_id,part_desc,qty) VALUES (?,?,?)",
                        (task_id, p['desc'], p.get('qty', 1))
                    )
                    conn.execute(
                        "INSERT INTO reorder (part_desc,qty,source) VALUES (?,?,?)",
                        (p['desc'], p.get('qty', 1), f'PM Task #{task_id}')
                    )
            conn.commit()
        log_action(d.get('completed_by_name', 'Artisan'), 'task_completed', f"Task #{task_id} completed")
        return jsonify({'status': 'ok'})
    except Exception as e:
        logger.error(f"Error completing task {task_id}: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': 'Failed to complete task'}), 500


@app.route('/api/tasks/<int:task_id>/checklist', methods=['POST'])
def update_task_checklist(task_id):
    """Update a single checklist item's checked state."""
    d = safe_json(request)
    try:
        with get_db_context() as conn:
            item_id = d.get('item_id')
            if item_id is not None:
                conn.execute(
                    "UPDATE pm_checklist_items SET checked=? WHERE id=?",
                    (1 if d.get('checked') else 0, int(item_id))
                )
                conn.commit()
        return jsonify({'status': 'ok'})
    except Exception as e:
        logger.error(f"Error updating checklist for task {task_id}: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': 'Failed to update checklist'}), 500


@app.route('/api/tasks/<int:task_id>/suspend', methods=['POST'])
def suspend_task(task_id):
    """Artisan suspends a job with a reason -> goes to admin 'Pushed' folder."""
    d = safe_json(request)
    reason = str(d.get('reason', '')).strip()
    if not reason:
        return jsonify({'status': 'error', 'message': 'A reason is required to suspend a job'}), 400
    try:
        with get_db_context() as conn:
            conn.execute("""
                UPDATE pm_tasks
                SET suspended=1, suspend_reason=?, suspended_at=?
                WHERE id=?
            """, (reason, datetime.now().isoformat(), task_id))
            conn.commit()
        log_action(d.get('by_name', 'Artisan'), 'task_suspended', f"Task #{task_id}: {reason}")
        return jsonify({'status': 'ok'})
    except Exception as e:
        logger.error(f"Error suspending task {task_id}: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': 'Failed to suspend task'}), 500


@app.route('/api/tasks/<int:task_id>/unsuspend', methods=['POST'])
def unsuspend_task(task_id):
    """Admin pushes a suspended job back to the artisan with a note."""
    d = safe_json(request)
    try:
        with get_db_context() as conn:
            conn.execute("""
                UPDATE pm_tasks
                SET suspended=0, suspend_reason=COALESCE(?, suspend_reason)
                WHERE id=?
            """, (d.get('note'), task_id))
            conn.commit()
        log_action('Admin', 'task_unsuspended', f"Task #{task_id} returned to artisan")
        return jsonify({'status': 'ok'})
    except Exception as e:
        logger.error(f"Error unsuspending task {task_id}: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': 'Failed to unsuspend task'}), 500


@app.route('/api/tasks/suspended')
def get_suspended_tasks():
    """Admin view: all suspended/pushed jobs."""
    try:
        with get_db_context() as conn:
            rows = conn.execute("""
                SELECT t.*, e.name as equip_name, u.name as assigned_to_name
                FROM pm_tasks t
                LEFT JOIN equipment e ON e.id = t.equipment_id
                LEFT JOIN users u ON u.id = t.assigned_to
                WHERE COALESCE(t.suspended,0)=1 AND t.status!='done'
                ORDER BY t.suspended_at DESC
            """).fetchall()
        return jsonify([dict(r) for r in rows])
    except Exception as e:
        logger.error(f"Error fetching suspended tasks: {e}", exc_info=True)
        return jsonify([]), 500


@app.route('/api/tasks/history/<int:user_id>')
def get_task_history(user_id):
    """DidIt tab: completed tasks for an artisan."""
    try:
        with get_db_context() as conn:
            rows = conn.execute("""
                SELECT t.*, e.name as equip_name
                FROM pm_tasks t
                LEFT JOIN equipment e ON e.id = t.equipment_id
                WHERE t.completed_by=? AND t.status='done'
                ORDER BY t.completed_at DESC LIMIT 100
            """, (user_id,)).fetchall()
        return jsonify([dict(r) for r in rows])
    except Exception as e:
        logger.error(f"Error fetching task history: {e}", exc_info=True)
        return jsonify([]), 500


# ============================================================
# BREAKDOWNS
# ============================================================
def _recalc_downtime(b):
    """Compute downtime minutes from start/end strings (HH:MM)."""
    def mins(t):
        try:
            h, m = str(t).split(':')[:2]
            return int(h) * 60 + int(m)
        except Exception:
            return None
    start = b.get('final_time_start') or b.get('art_time_start') or b.get('sup_time_start')
    end = b.get('final_time_end') or b.get('art_time_end') or b.get('sup_time_end')
    s, e = mins(start), mins(end)
    if s is None or e is None:
        return None
    dur = e - s
    if dur < 0:
        dur += 1440  # crossed midnight
    return dur


@app.route('/api/breakdowns/supervisor', methods=['POST'])
def log_supervisor_breakdown():
    """Supervisor/manager logs a breakdown -> appears on artisan board."""
    d = safe_json(request)
    if not d.get('equipment_id') or not d.get('artisan_id'):
        return jsonify({'status': 'error', 'message': 'equipment_id and artisan_id required'}), 400
    try:
        with get_db_context() as conn:
            auto_time = datetime.now().strftime('%H:%M')
            cur = conn.execute("""
                INSERT INTO breakdowns
                (log_date, supervisor_id, artisan_id, equipment_id,
                 sup_time_start, sup_time_end, sup_failure_type,
                 sup_component, sup_failure_mode, sup_notes, sup_machine_status,
                 status, source)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,'supervisor_logged','live')
            """, (
                date.today().isoformat(), d.get('supervisor_id'), d.get('artisan_id'),
                d.get('equipment_id'), d.get('time_start') or auto_time, d.get('time_end'),
                d.get('failure_type'), d.get('component'), d.get('failure_mode'),
                d.get('supervisor_notes'), d.get('machine_status')
            ))
            bd_id = cur.lastrowid
            conn.commit()
        log_action('Supervisor', 'breakdown_logged', f"Breakdown #{bd_id} -> artisan {d.get('artisan_id')}")
        return jsonify({'status': 'ok', 'id': bd_id})
    except Exception as e:
        logger.error(f"Error logging breakdown: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': 'Failed to log breakdown'}), 500


@app.route('/api/breakdowns')
def get_breakdowns():
    """List breakdowns within N days, with names + parts."""
    days = get_days_arg(14)
    try:
        cutoff = (date.today() - timedelta(days=days)).isoformat()
        with get_db_context() as conn:
            rows = conn.execute("""
                SELECT b.*, e.name as equip_name, e.area,
                       a.name as artisan_name, s.name as supervisor_name
                FROM breakdowns b
                LEFT JOIN equipment e ON e.id = b.equipment_id
                LEFT JOIN users a ON a.id = b.artisan_id
                LEFT JOIN users s ON s.id = b.supervisor_id
                WHERE b.log_date >= ?
                ORDER BY b.logged_at DESC
            """, (cutoff,)).fetchall()
            result = []
            for b in rows:
                bd = dict(b)
                parts = conn.execute(
                    "SELECT part_desc, qty FROM breakdown_parts WHERE breakdown_id=?",
                    (b['id'],)
                ).fetchall()
                bd['parts_list'] = ', '.join(f"{p['part_desc']} x{p['qty']}" for p in parts) if parts else ''
                result.append(bd)
        return jsonify(result)
    except Exception as e:
        logger.error(f"Error fetching breakdowns: {e}", exc_info=True)
        return jsonify([]), 500


@app.route('/api/breakdowns/supervisor')
def get_supervisor_breakdowns():
    """Open breakdowns (awaiting artisan)."""
    try:
        with get_db_context() as conn:
            rows = conn.execute("""
                SELECT b.*, e.name as equip_name, e.area,
                       a.name as artisan_name
                FROM breakdowns b
                LEFT JOIN equipment e ON e.id = b.equipment_id
                LEFT JOIN users a ON a.id = b.artisan_id
                WHERE b.status='supervisor_logged'
                ORDER BY b.logged_at DESC
            """).fetchall()
        return jsonify([dict(r) for r in rows])
    except Exception as e:
        logger.error(f"Error fetching supervisor breakdowns: {e}", exc_info=True)
        return jsonify([]), 500


@app.route('/api/breakdowns/mine/<int:user_id>')
def get_my_breakdowns(user_id):
    """An artisan's breakdowns."""
    days = get_days_arg(14)
    try:
        cutoff = (date.today() - timedelta(days=days)).isoformat()
        with get_db_context() as conn:
            rows = conn.execute("""
                SELECT b.*, e.name as equip_name, e.area,
                       a.name as artisan_name
                FROM breakdowns b
                LEFT JOIN equipment e ON e.id = b.equipment_id
                LEFT JOIN users a ON a.id = b.artisan_id
                WHERE b.artisan_id=? AND b.log_date >= ?
                ORDER BY b.logged_at DESC
            """, (user_id, cutoff)).fetchall()
        return jsonify([dict(r) for r in rows])
    except Exception as e:
        logger.error(f"Error fetching artisan breakdowns: {e}", exc_info=True)
        return jsonify([]), 500


@app.route('/api/breakdowns/disputed')
def get_disputed_breakdowns():
    """Disputed breakdowns for admin resolution."""
    try:
        with get_db_context() as conn:
            rows = conn.execute("""
                SELECT b.*, e.name as equip_name,
                       a.name as artisan_name, s.name as supervisor_name
                FROM breakdowns b
                LEFT JOIN equipment e ON e.id = b.equipment_id
                LEFT JOIN users a ON a.id = b.artisan_id
                LEFT JOIN users s ON s.id = b.supervisor_id
                WHERE b.disputed=1 AND b.status != 'complete'
                ORDER BY b.logged_at DESC
            """).fetchall()
        return jsonify([dict(r) for r in rows])
    except Exception as e:
        logger.error(f"Error fetching disputed breakdowns: {e}", exc_info=True)
        return jsonify([]), 500


@app.route('/api/breakdowns/<int:bd_id>/artisan', methods=['POST'])
def artisan_complete_breakdown(bd_id):
    """Artisan files their breakdown report.
    If supervisor hasn't completed yet → artisan_completed (waiting for supervisor).
    If supervisor already completed → compare times, set complete or disputed."""
    d = safe_json(request)
    try:
        with get_db_context() as conn:
            row = conn.execute("SELECT * FROM breakdowns WHERE id=?", (bd_id,)).fetchone()
            if not row:
                return jsonify({'status': 'error', 'message': 'Breakdown not found'}), 404
            b = dict(row)
            sup_done = b.get('status') == 'awaiting_artisan_review'

            disputed = 0
            dispute_field = None
            if sup_done:
                # Compare artisan times against supervisor's recorded times
                art_start = d.get('art_time_start', '')
                art_end   = d.get('art_time_end', '')
                sup_start = (b.get('sup_time_start') or '')[:5]
                sup_end   = (b.get('sup_time_end') or '')[:5]
                if not sup_end:
                    ct = b.get('sup_completed_at') or ''
                    sup_end = ct[11:16] if len(ct) > 15 else ''
                if art_start and sup_start and art_start != sup_start:
                    disputed = 1; dispute_field = 'time_start'
                if art_end and sup_end and art_end != sup_end:
                    disputed = 1
                    dispute_field = (dispute_field + ', time_end') if dispute_field else 'time_end'

            b['art_time_start'] = d.get('art_time_start')
            b['art_time_end']   = d.get('art_time_end')
            downtime = _recalc_downtime(b)
            ftype = d.get('failure_type') or b.get('sup_failure_type')
            new_status = ('disputed' if disputed else 'complete') if sup_done else 'artisan_completed'

            conn.execute("""
                UPDATE breakdowns SET
                    art_time_start=?, art_time_end=?, art_failure_type=?,
                    art_description=?, art_repair_action=?, art_machine_status=?,
                    art_followup=?, art_failed_component=?, art_failure_mode=?,
                    art_completed_at=?, status=?, disputed=?, dispute_field=?,
                    downtime_mins=?,
                    production_downtime_mins=CASE WHEN ?='Production' THEN ? ELSE 0 END,
                    maintenance_downtime_mins=CASE WHEN ?='Production' THEN 0 ELSE ? END
                WHERE id=?
            """, (
                d.get('art_time_start'), d.get('art_time_end'), d.get('failure_type'),
                d.get('description'), d.get('repair_action'), d.get('machine_status'),
                d.get('followup'), d.get('failed_component'), d.get('failure_mode'),
                datetime.now().isoformat(), new_status, disputed, dispute_field,
                downtime, ftype, downtime, ftype, downtime, bd_id
            ))
            for p in (d.get('parts', []) or []):
                if p.get('desc'):
                    conn.execute(
                        "INSERT INTO breakdown_parts (breakdown_id,part_desc,qty,logged_by) VALUES (?,?,?,'artisan')",
                        (bd_id, p['desc'], p.get('qty', 1))
                    )
                    conn.execute(
                        "INSERT INTO reorder (part_desc,qty,source) VALUES (?,?,?)",
                        (p['desc'], p.get('qty', 1), f'Breakdown #{bd_id}')
                    )
            conn.commit()
        log_action(d.get('artisan_name', 'Artisan'), 'breakdown_completed', f"Breakdown #{bd_id} ({new_status})")
        return jsonify({'status': 'ok', 'disputed': bool(disputed), 'new_status': new_status})
    except Exception as e:
        logger.error(f"Error completing breakdown {bd_id}: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': 'Failed to complete breakdown'}), 500


@app.route('/api/breakdowns/<int:bd_id>/resolve', methods=['POST'])
def resolve_breakdown(bd_id):
    """Admin resolves a disputed breakdown with final values and optional dept allocations."""
    d = safe_json(request)
    try:
        with get_db_context() as conn:
            b = dict(conn.execute("SELECT * FROM breakdowns WHERE id=?", (bd_id,)).fetchone() or {})
            b['final_time_start'] = d.get('final_time_start')
            b['final_time_end']   = d.get('final_time_end')
            downtime = _recalc_downtime(b)
            ftype = d.get('final_failure_type') or b.get('sup_failure_type')
            conn.execute("""
                UPDATE breakdowns SET
                    final_time_start=?, final_time_end=?, final_description=?,
                    final_failure_type=?, admin_resolved_by=?, admin_resolution_note=?,
                    admin_resolved_at=?, status='complete', disputed=0,
                    downtime_mins=?,
                    production_downtime_mins=CASE WHEN ?='Production' THEN ? ELSE 0 END,
                    maintenance_downtime_mins=CASE WHEN ?='Production' THEN 0 ELSE ? END
                WHERE id=?
            """, (
                d.get('final_time_start'), d.get('final_time_end'), d.get('final_description'),
                ftype, d.get('admin_id'), d.get('resolution_note'),
                datetime.now().isoformat(), downtime,
                ftype, downtime, ftype, downtime, bd_id
            ))
            # Save dept allocations — replace any existing ones for this breakdown
            conn.execute("DELETE FROM breakdown_allocations WHERE breakdown_id=?", (bd_id,))
            for alloc in (d.get('allocations') or []):
                mins = int(alloc.get('mins') or 0)
                if alloc.get('dept') and mins > 0:
                    conn.execute(
                        "INSERT INTO breakdown_allocations (breakdown_id, dept, mins, note) VALUES (?,?,?,?)",
                        (bd_id, alloc['dept'], mins, alloc.get('note', ''))
                    )
            conn.commit()
        log_action('Admin', 'breakdown_resolved', f"Breakdown #{bd_id} resolved")
        return jsonify({'status': 'ok'})
    except Exception as e:
        logger.error(f"Error resolving breakdown {bd_id}: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': 'Failed to resolve breakdown'}), 500


@app.route('/api/breakdowns/<int:bd_id>/direct-order', methods=['POST'])
def breakdown_direct_order(bd_id):
    """Direct parts order from a breakdown (Drymill flow) -> reorder list."""
    d = safe_json(request)
    try:
        with get_db_context() as conn:
            for p in (d.get('parts', []) or []):
                if p.get('desc'):
                    conn.execute(
                        "INSERT INTO breakdown_parts (breakdown_id,part_desc,qty,logged_by) VALUES (?,?,?,'supervisor')",
                        (bd_id, p['desc'], p.get('qty', 1))
                    )
                    conn.execute(
                        "INSERT INTO reorder (part_desc,qty,source) VALUES (?,?,?)",
                        (p['desc'], p.get('qty', 1), f'Direct order BD#{bd_id}')
                    )
            conn.commit()
        return jsonify({'status': 'ok'})
    except Exception as e:
        logger.error(f"Error in direct order for breakdown {bd_id}: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': 'Failed to create direct order'}), 500


# ============================================================
# JOB REQUESTS  (supervisor/manager -> admin)
# ============================================================
@app.route('/api/jobrequests')
def get_job_requests():
    try:
        with get_db_context() as conn:
            rows = conn.execute("""
                SELECT j.*, e.name as equip_name, u.name as requester_name
                FROM job_requests j
                LEFT JOIN equipment e ON e.id = j.equipment_id
                LEFT JOIN users u ON u.id = j.requested_by
                ORDER BY j.created_at DESC
            """).fetchall()
        return jsonify([dict(r) for r in rows])
    except Exception as e:
        logger.error(f"Error fetching job requests: {e}", exc_info=True)
        return jsonify([]), 500


@app.route('/api/jobrequests', methods=['POST'])
def add_job_request():
    d = safe_json(request)
    if not d.get('description'):
        return jsonify({'status': 'error', 'message': 'description is required'}), 400
    try:
        with get_db_context() as conn:
            cur = conn.execute("""
                INSERT INTO job_requests (requested_by, equipment_id, description, urgency, status)
                VALUES (?,?,?,?,'pending')
            """, (d.get('requested_by'), d.get('equipment_id'), d['description'], d.get('urgency', 'Normal')))
            jid = cur.lastrowid
            conn.commit()
        log_action(d.get('requester_name', 'User'), 'job_requested', f"Job request #{jid}")
        return jsonify({'status': 'ok', 'id': jid})
    except Exception as e:
        logger.error(f"Error adding job request: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': 'Failed to add job request'}), 500


@app.route('/api/jobrequests/<int:jid>', methods=['PUT'])
def update_job_request(jid):
    d = safe_json(request)
    try:
        with get_db_context() as conn:
            conn.execute("""
                UPDATE job_requests SET status=?, admin_note=?, resolved_at=?
                WHERE id=?
            """, (d.get('status', 'pending'), d.get('admin_note', ''), datetime.now().isoformat(), jid))
            conn.commit()
        return jsonify({'status': 'ok'})
    except Exception as e:
        logger.error(f"Error updating job request {jid}: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': 'Failed to update job request'}), 500


# ============================================================
# DASHBOARD (weekly aggregate for admin/manager/supervisor)
# ============================================================
@app.route('/api/dashboard/weekly')
def dashboard_weekly():
    try:
        week_start = _week_start_iso()
        cutoff = (date.today() - timedelta(days=7)).isoformat()
        with get_db_context() as conn:
            # PM compliance by area
            compliance = conn.execute("""
                SELECT COALESCE(e.area,'General') as name,
                       COUNT(*) as total,
                       SUM(CASE WHEN t.status='done' THEN 1 ELSE 0 END) as done
                FROM pm_tasks t
                LEFT JOIN equipment e ON e.id = t.equipment_id
                WHERE t.week_start=?
                GROUP BY e.area
            """, (week_start,)).fetchall()

            # Machines by downtime (this week)
            machines = conn.execute("""
                SELECT e.name, e.area,
                       COUNT(*) as failures,
                       SUM(COALESCE(b.downtime_mins,0)) as total_mins,
                       MAX(b.sup_failure_type) as sup_failure_type
                FROM breakdowns b
                LEFT JOIN equipment e ON e.id = b.equipment_id
                WHERE b.log_date >= ?
                GROUP BY b.equipment_id
                ORDER BY total_mins DESC
                LIMIT 10
            """, (cutoff,)).fetchall()

            open_bd = conn.execute(
                "SELECT COUNT(*) FROM breakdowns WHERE status='supervisor_logged'"
            ).fetchone()[0]
            disputed = conn.execute(
                "SELECT COUNT(*) FROM breakdowns WHERE disputed=1 AND status!='complete'"
            ).fetchone()[0]
            pending_req = conn.execute(
                "SELECT COUNT(*) FROM job_requests WHERE status='pending'"
            ).fetchone()[0]

            # PM gauges: day / week / month completion %
            today = date.today().isoformat()
            month_start = date.today().replace(day=1).isoformat()

            def pct(where, params):
                r = conn.execute(
                    f"SELECT COUNT(*) total, SUM(CASE WHEN status='done' THEN 1 ELSE 0 END) done "
                    f"FROM pm_tasks WHERE {where}", params
                ).fetchone()
                total = r['total'] or 0
                done = r['done'] or 0
                return round(done / total * 100) if total else 0, done, total

            day_pct, day_done, day_total = pct("due_date=?", (today,))
            week_pct, week_done, week_total = pct("week_start=?", (week_start,))
            month_pct, month_done, month_total = pct("created_at >= ?", (month_start,))

        return jsonify({
            'compliance': [dict(c) for c in compliance],
            'machines': [dict(m) for m in machines],
            'open_breakdowns': open_bd,
            'disputed': disputed,
            'pending_requests': pending_req,
            'pending_request': pending_req,
            'gauges': {
                'day': {'pct': day_pct, 'done': day_done, 'total': day_total},
                'week': {'pct': week_pct, 'done': week_done, 'total': week_total},
                'month': {'pct': month_pct, 'done': month_done, 'total': month_total},
            }
        })
    except Exception as e:
        logger.error(f"Error building dashboard: {e}", exc_info=True)
        return jsonify({}), 500


# ============================================================
# ACTIVITY LOG
# ============================================================
@app.route('/api/log')
def get_activity_log():
    days = get_days_arg(7)
    try:
        cutoff = (datetime.now() - timedelta(days=days)).isoformat()
        with get_db_context() as conn:
            rows = conn.execute("""
                SELECT * FROM activity_log
                WHERE logged_at >= ?
                ORDER BY logged_at DESC LIMIT 200
            """, (cutoff,)).fetchall()
        return jsonify([dict(r) for r in rows])
    except Exception as e:
        logger.error(f"Error fetching activity log: {e}", exc_info=True)
        return jsonify([]), 500


# ============================================================
# SCHEDULE (recurring PM)
# ============================================================
@app.route('/api/schedule')
def get_schedule():
    try:
        with get_db_context() as conn:
            rows = conn.execute("""
                SELECT s.*, e.name as equip_name, u.name as assigned_to_name
                FROM pm_schedule s
                LEFT JOIN equipment e ON e.id = s.equipment_id
                LEFT JOIN users u ON u.id = s.assigned_to
                ORDER BY s.title
            """).fetchall()
        return jsonify([dict(r) for r in rows])
    except Exception as e:
        logger.error(f"Error fetching schedule: {e}", exc_info=True)
        return jsonify([]), 500


@app.route('/api/schedule', methods=['POST'])
def add_schedule():
    d = safe_json(request)
    if not d.get('title'):
        return jsonify({'status': 'error', 'message': 'title required'}), 400
    try:
        with get_db_context() as conn:
            conn.execute("""
                INSERT INTO pm_schedule (title, equipment_id, assigned_to, frequency, shift, active)
                VALUES (?,?,?,?,?,1)
            """, (d['title'], d.get('equipment_id'), d.get('assigned_to'),
                  d.get('frequency', 'weekly'), d.get('shift', 'day')))
            conn.commit()
        return jsonify({'status': 'ok'})
    except Exception as e:
        logger.error(f"Error adding schedule: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': 'Failed to add schedule'}), 500


@app.route('/api/schedule/<int:sid>/toggle', methods=['POST'])
def toggle_schedule(sid):
    try:
        with get_db_context() as conn:
            conn.execute("UPDATE pm_schedule SET active = CASE WHEN active=1 THEN 0 ELSE 1 END WHERE id=?", (sid,))
            conn.commit()
        return jsonify({'status': 'ok'})
    except Exception as e:
        logger.error(f"Error toggling schedule {sid}: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': 'Failed to toggle'}), 500


@app.route('/api/schedule/<int:sid>', methods=['DELETE'])
def delete_schedule(sid):
    try:
        with get_db_context() as conn:
            conn.execute("DELETE FROM pm_schedule WHERE id=?", (sid,))
            conn.commit()
        return jsonify({'status': 'ok'})
    except Exception as e:
        logger.error(f"Error deleting schedule {sid}: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': 'Failed to delete'}), 500


# ============================================================
# TEMPLATES (checklist / order templates)
# ============================================================
@app.route('/api/templates')
def get_templates():
    """Return checklist templates as {name: [{item}, ...]}."""
    try:
        with get_db_context() as conn:
            rows = conn.execute(
                "SELECT task_type, item, sort_order FROM checklist_templates ORDER BY task_type, sort_order"
            ).fetchall()
        out = {}
        for r in rows:
            out.setdefault(r['task_type'], []).append({'item': r['item']})
        return jsonify(out)
    except Exception as e:
        logger.error(f"Error fetching templates: {e}", exc_info=True)
        return jsonify({}), 500


@app.route('/api/templates/<name>', methods=['POST'])
def save_template(name):
    d = safe_json(request)
    items = d.get('items', [])
    try:
        with get_db_context() as conn:
            conn.execute("DELETE FROM checklist_templates WHERE task_type=?", (name,))
            for i, item in enumerate(items):
                txt = item.get('item') if isinstance(item, dict) else item
                if txt:
                    conn.execute(
                        "INSERT INTO checklist_templates (task_type,item,sort_order) VALUES (?,?,?)",
                        (name, txt, i)
                    )
            conn.commit()
        return jsonify({'status': 'ok'})
    except Exception as e:
        logger.error(f"Error saving template {name}: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': 'Failed to save template'}), 500


@app.route('/api/templates/<name>', methods=['DELETE'])
def delete_template(name):
    try:
        with get_db_context() as conn:
            conn.execute("DELETE FROM checklist_templates WHERE task_type=?", (name,))
            conn.commit()
        return jsonify({'status': 'ok'})
    except Exception as e:
        logger.error(f"Error deleting template {name}: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': 'Failed to delete template'}), 500


# ============================================================
# READINGS (boiler / utility)
# ============================================================
@app.route('/api/readings/boiler', methods=['GET', 'POST'])
def boiler_readings():
    if request.method == 'POST':
        d = safe_json(request)
        try:
            with get_db_context() as conn:
                conn.execute("""
                    INSERT INTO boiler_readings
                    (log_date, boiler_id, logged_by, tds_morning, tds_night,
                     blowdowns_total, condensate_meter_reading, makeup_meter_reading,
                     softener_day, softener_night, salt_bags, lost_time_mins, notes)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
                """, (
                    d.get('log_date', date.today().isoformat()), d.get('boiler_id'),
                    d.get('logged_by'), d.get('tds_morning'), d.get('tds_night'),
                    d.get('blowdowns_total', 0), d.get('condensate_meter_reading'),
                    d.get('makeup_meter_reading'), d.get('softener_day'),
                    d.get('softener_night'), d.get('salt_bags'),
                    d.get('lost_time_mins'), d.get('notes')
                ))
                conn.commit()
            return jsonify({'status': 'ok'})
        except Exception as e:
            logger.error(f"Error saving boiler reading: {e}", exc_info=True)
            return jsonify({'status': 'error', 'message': 'Failed to save reading'}), 500
    # GET
    days = get_days_arg(7)
    try:
        cutoff = (date.today() - timedelta(days=days)).isoformat()
        with get_db_context() as conn:
            rows = conn.execute("""
                SELECT br.*, e.name as boiler_name
                FROM boiler_readings br
                LEFT JOIN equipment e ON e.id = br.boiler_id
                WHERE br.log_date >= ?
                ORDER BY br.log_date DESC
            """, (cutoff,)).fetchall()
        return jsonify([dict(r) for r in rows])
    except Exception as e:
        logger.error(f"Error fetching boiler readings: {e}", exc_info=True)
        return jsonify([]), 500


@app.route('/api/readings/utility', methods=['GET', 'POST'])
def utility_readings():
    if request.method == 'POST':
        d = safe_json(request)
        try:
            with get_db_context() as conn:
                conn.execute("""
                    INSERT INTO utility_readings
                    (log_date, substation_id, hour_meter, run_hours_today,
                     fuel_level, diesel_issued, oil_level, status, notes)
                    VALUES (?,?,?,?,?,?,?,?,?)
                """, (
                    d.get('log_date', date.today().isoformat()), d.get('substation_id'),
                    d.get('hour_meter'), d.get('run_hours_today'), d.get('fuel_level'),
                    d.get('diesel_issued', 0), d.get('oil_level'),
                    d.get('status', 'standby'), d.get('notes')
                ))
                conn.commit()
            return jsonify({'status': 'ok'})
        except Exception as e:
            logger.error(f"Error saving utility reading: {e}", exc_info=True)
            return jsonify({'status': 'error', 'message': 'Failed to save reading'}), 500
    # GET
    days = get_days_arg(7)
    try:
        cutoff = (date.today() - timedelta(days=days)).isoformat()
        with get_db_context() as conn:
            rows = conn.execute("""
                SELECT ur.*, e.name as equip_name
                FROM utility_readings ur
                LEFT JOIN equipment e ON e.id = ur.substation_id
                WHERE ur.log_date >= ?
                ORDER BY ur.log_date DESC
            """, (cutoff,)).fetchall()
        return jsonify([dict(r) for r in rows])
    except Exception as e:
        logger.error(f"Error fetching utility readings: {e}", exc_info=True)
        return jsonify([]), 500


# ============================================================
# REORDER LIST
# ============================================================
@app.route('/api/reorder')
def get_reorder():
    try:
        with get_db_context() as conn:
            rows = conn.execute(
                "SELECT * FROM reorder WHERE cleared=0 ORDER BY logged_at DESC"
            ).fetchall()
        return jsonify([dict(r) for r in rows])
    except Exception as e:
        logger.error(f"Error fetching reorder list: {e}", exc_info=True)
        return jsonify([]), 500


@app.route('/api/reorder', methods=['POST'])
def add_reorder():
    d = safe_json(request)
    if not d.get('part_desc'):
        return jsonify({'status': 'error', 'message': 'part_desc required'}), 400
    try:
        with get_db_context() as conn:
            conn.execute(
                "INSERT INTO reorder (part_desc,qty,source) VALUES (?,?,?)",
                (d['part_desc'], d.get('qty', 1), d.get('source', 'Manual'))
            )
            conn.commit()
        return jsonify({'status': 'ok'})
    except Exception as e:
        logger.error(f"Error adding reorder: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': 'Failed to add reorder item'}), 500


@app.route('/api/reorder/clear', methods=['POST'])
def clear_reorder():
    try:
        with get_db_context() as conn:
            conn.execute("UPDATE reorder SET cleared=1 WHERE cleared=0")
            conn.commit()
        return jsonify({'status': 'ok'})
    except Exception as e:
        logger.error(f"Error clearing reorder: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': 'Failed to clear reorder'}), 500


# ============================================================
# STORES — Stock numbering, stock, movements, dept codes
# ============================================================
def _next_stock_code(conn, dept_code):
    """Generate next stock code: SITE-PREFIX-DEPT-NNNNNN."""
    site = get_setting('stores_site_code', 'NW')
    prefix = get_setting('stores_prefix', 'STK')
    dept = (dept_code or 'GEN').upper()
    like = f"{site}-{prefix}-{dept}-%"
    row = conn.execute(
        "SELECT stock_code FROM stores_stock WHERE stock_code LIKE ? ORDER BY id DESC LIMIT 1",
        (like,)
    ).fetchone()
    seq = 1
    if row and row['stock_code']:
        try:
            seq = int(row['stock_code'].split('-')[-1]) + 1
        except (ValueError, IndexError):
            seq = 1
    return f"{site}-{prefix}-{dept}-{seq:06d}"


@app.route('/api/stores/stock')
def get_stock():
    try:
        with get_db_context() as conn:
            rows = conn.execute("""
                SELECT s.*,
                       CASE WHEN s.qty_on_hand <= s.qty_min_stock THEN 1 ELSE 0 END as low_stock
                FROM stores_stock s
                WHERE s.active=1
                ORDER BY s.is_consumable DESC, s.part_desc
            """).fetchall()
        return jsonify([dict(r) for r in rows])
    except Exception as e:
        logger.error(f"Error fetching stock: {e}", exc_info=True)
        return jsonify([]), 500


@app.route('/api/stores/stock/add', methods=['POST'])
def add_stock():
    d = safe_json(request)
    if not d.get('part_desc'):
        return jsonify({'status': 'error', 'message': 'part_desc required'}), 400
    try:
        with get_db_context() as conn:
            code = _next_stock_code(conn, d.get('dept_code'))
            conn.execute("""
                INSERT INTO stores_stock
                (stock_code, part_desc, part_number, dept_code, category, unit,
                 qty_on_hand, qty_min_stock, qty_max_stock, bin_number, location,
                 supplier, is_consumable, is_critical, critical_equipment_id,
                 created_by)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, (
                code, d['part_desc'], d.get('part_number'), d.get('dept_code'),
                d.get('category', 'General'), d.get('unit', 'each'),
                d.get('qty_on_hand', 0), d.get('qty_min_stock', 0),
                d.get('qty_max_stock', 0), d.get('bin_number'), d.get('location'),
                d.get('supplier'), int(d.get('is_consumable', 0) or 0),
                int(d.get('is_critical', 0) or 0), d.get('critical_equipment_id'),
                d.get('created_by')
            ))
            sid = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
            # opening balance movement
            if d.get('qty_on_hand', 0):
                conn.execute("""
                    INSERT INTO stock_movements
                    (stock_id, movement_type, qty, qty_after, reference, logged_by_name)
                    VALUES (?,?,?,?,?,?)
                """, (sid, 'receipt', d.get('qty_on_hand', 0), d.get('qty_on_hand', 0),
                      'Opening balance', 'System'))
            conn.commit()
        return jsonify({'status': 'ok', 'stock_code': code, 'id': sid})
    except Exception as e:
        logger.error(f"Error adding stock: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': 'Failed to add stock'}), 500


@app.route('/api/stores/stock/<int:sid>', methods=['DELETE'])
def delete_stock(sid):
    try:
        with get_db_context() as conn:
            conn.execute("UPDATE stores_stock SET active=0 WHERE id=?", (sid,))
            conn.commit()
        return jsonify({'status': 'ok'})
    except Exception as e:
        logger.error(f"Error deleting stock {sid}: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': 'Failed to delete stock'}), 500


@app.route('/api/stores/movement', methods=['POST'])
def stock_movement():
    """Issue / receive / adjust / return stock."""
    d = safe_json(request)
    sid = d.get('stock_id')
    mtype = d.get('movement_type')
    qty = float(d.get('qty', 0) or 0)
    if not sid or not mtype or qty <= 0:
        return jsonify({'status': 'error', 'message': 'stock_id, movement_type and qty required'}), 400
    try:
        with get_db_context() as conn:
            row = conn.execute("SELECT qty_on_hand FROM stores_stock WHERE id=?", (sid,)).fetchone()
            if not row:
                return jsonify({'status': 'error', 'message': 'Stock item not found'}), 404
            on_hand = row['qty_on_hand']
            if mtype in ('issue',) and qty > on_hand:
                return jsonify({'status': 'error', 'message': f'Only {on_hand} on hand'}), 400
            delta = qty if mtype in ('receipt', 'return', 'adjustment') else -qty
            new_qty = on_hand + delta
            conn.execute("UPDATE stores_stock SET qty_on_hand=? WHERE id=?", (new_qty, sid))
            conn.execute("""
                INSERT INTO stock_movements
                (stock_id, movement_type, qty, qty_after, job_number, reference, notes,
                 logged_by, logged_by_name)
                VALUES (?,?,?,?,?,?,?,?,?)
            """, (sid, mtype, qty, new_qty, d.get('job_number'), d.get('reference'),
                  d.get('notes'), d.get('logged_by'), d.get('logged_by_name', 'Stores')))
            conn.commit()
        return jsonify({'status': 'ok', 'new_qty': new_qty})
    except Exception as e:
        logger.error(f"Error recording movement: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': 'Failed to record movement'}), 500


@app.route('/api/stores/movements')
def get_movements():
    days = get_days_arg(30)
    try:
        cutoff = (datetime.now() - timedelta(days=days)).isoformat()
        with get_db_context() as conn:
            rows = conn.execute("""
                SELECT m.*, s.part_desc, s.stock_code
                FROM stock_movements m
                LEFT JOIN stores_stock s ON s.id = m.stock_id
                WHERE m.logged_at >= ?
                ORDER BY m.logged_at DESC LIMIT 200
            """, (cutoff,)).fetchall()
        return jsonify([dict(r) for r in rows])
    except Exception as e:
        logger.error(f"Error fetching movements: {e}", exc_info=True)
        return jsonify([]), 500


@app.route('/api/stores/dept-codes')
def get_dept_codes():
    try:
        with get_db_context() as conn:
            rows = conn.execute(
                "SELECT code, label FROM dept_codes ORDER BY sort_order, code"
            ).fetchall()
        return jsonify([dict(r) for r in rows])
    except Exception as e:
        logger.error(f"Error fetching dept codes: {e}", exc_info=True)
        return jsonify([]), 500


@app.route('/api/stores/dept-codes', methods=['POST'])
def add_dept_code():
    d = safe_json(request)
    code = str(d.get('code', '')).strip().upper()
    label = str(d.get('label', '')).strip()
    if not code or not label:
        return jsonify({'status': 'error', 'message': 'code and label required'}), 400
    try:
        with get_db_context() as conn:
            existing = conn.execute("SELECT 1 FROM dept_codes WHERE code=?", (code,)).fetchone()
            if existing:
                return jsonify({'status': 'error', 'message': 'Code already exists'}), 400
            conn.execute("INSERT INTO dept_codes (code,label) VALUES (?,?)", (code, label))
            conn.commit()
        return jsonify({'status': 'ok'})
    except Exception as e:
        logger.error(f"Error adding dept code: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': 'Failed to add dept code'}), 500


@app.route('/api/stores/dept-codes/<code>', methods=['DELETE'])
def delete_dept_code(code):
    try:
        with get_db_context() as conn:
            conn.execute("DELETE FROM dept_codes WHERE code=?", (code.upper(),))
            conn.commit()
        return jsonify({'status': 'ok'})
    except Exception as e:
        logger.error(f"Error deleting dept code {code}: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': 'Failed to delete dept code'}), 500


@app.route('/api/stores/settings', methods=['GET', 'POST'])
def stores_settings():
    if request.method == 'POST':
        d = safe_json(request)
        try:
            if 'stores_site_code' in d:
                set_setting('stores_site_code', str(d['stores_site_code']).upper())
            if 'stores_prefix' in d:
                set_setting('stores_prefix', str(d['stores_prefix']).upper())
            return jsonify({'status': 'ok'})
        except Exception as e:
            logger.error(f"Error saving stores settings: {e}", exc_info=True)
            return jsonify({'status': 'error', 'message': 'Failed to save settings'}), 500
    return jsonify({
        'stores_site_code': get_setting('stores_site_code', 'NW'),
        'stores_prefix': get_setting('stores_prefix', 'STK'),
    })


@app.route('/api/stores/import-reorder', methods=['POST'])
def import_reorder_to_stores():
    """Create stock entries from outstanding reorder items (no duplicates)."""
    try:
        added = 0
        with get_db_context() as conn:
            items = conn.execute(
                "SELECT DISTINCT part_desc FROM reorder WHERE cleared=0"
            ).fetchall()
            for it in items:
                desc = it['part_desc']
                exists = conn.execute(
                    "SELECT 1 FROM stores_stock WHERE part_desc=? AND active=1", (desc,)
                ).fetchone()
                if not exists:
                    code = _next_stock_code(conn, 'GEN')
                    conn.execute("""
                        INSERT INTO stores_stock (stock_code, part_desc, dept_code, category, unit, qty_on_hand)
                        VALUES (?,?,?,?,?,0)
                    """, (code, desc, 'GEN', 'General', 'each'))
                    added += 1
            conn.commit()
        return jsonify({'status': 'ok', 'added': added})
    except Exception as e:
        logger.error(f"Error importing reorder: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': 'Failed to import'}), 500


# ============================================================
# EQUIPMENT SPARES
# ============================================================
@app.route('/api/equipment/<int:equip_id>/spares')
def get_equipment_spares(equip_id):
    try:
        with get_db_context() as conn:
            rows = conn.execute(
                "SELECT * FROM equipment_spares WHERE equipment_id=? ORDER BY part_desc",
                (equip_id,)
            ).fetchall()
        return jsonify([dict(r) for r in rows])
    except Exception as e:
        logger.error(f"Error fetching spares: {e}", exc_info=True)
        return jsonify([]), 500


@app.route('/api/equipment/<int:equip_id>/spares', methods=['POST'])
def add_equipment_spare(equip_id):
    d = safe_json(request)
    if not d.get('part_desc'):
        return jsonify({'status': 'error', 'message': 'part_desc required'}), 400
    try:
        with get_db_context() as conn:
            conn.execute("""
                INSERT INTO equipment_spares
                (equipment_id, part_desc, part_number, std_qty, unit, category, created_by)
                VALUES (?,?,?,?,?,?,?)
            """, (equip_id, d['part_desc'], d.get('part_number'),
                  d.get('std_qty', 1), d.get('unit', 'each'),
                  d.get('category', 'General'), d.get('created_by')))
            conn.commit()
        return jsonify({'status': 'ok'})
    except Exception as e:
        logger.error(f"Error adding spare: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': 'Failed to add spare'}), 500


@app.route('/api/equipment/spares/<int:spare_id>', methods=['DELETE'])
def delete_equipment_spare(spare_id):
    try:
        with get_db_context() as conn:
            conn.execute("DELETE FROM equipment_spares WHERE id=?", (spare_id,))
            conn.commit()
        return jsonify({'status': 'ok'})
    except Exception as e:
        logger.error(f"Error deleting spare {spare_id}: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': 'Failed to delete spare'}), 500


# ============================================================
# SUPPLIERS
# ============================================================
@app.route('/api/suppliers')
def get_suppliers():
    try:
        with get_db_context() as conn:
            rows = conn.execute(
                "SELECT * FROM suppliers WHERE active=1 ORDER BY category, name"
            ).fetchall()
        return jsonify([dict(r) for r in rows])
    except Exception as e:
        logger.error(f"Error fetching suppliers: {e}", exc_info=True)
        return jsonify([]), 500


@app.route('/api/suppliers', methods=['POST'])
def add_supplier():
    d = safe_json(request)
    if not d.get('name'):
        return jsonify({'status': 'error', 'message': 'name required'}), 400
    try:
        with get_db_context() as conn:
            conn.execute(
                "INSERT INTO suppliers (name,email,category) VALUES (?,?,?)",
                (d['name'], d.get('email'), d.get('category', 'General'))
            )
            conn.commit()
        return jsonify({'status': 'ok'})
    except Exception as e:
        logger.error(f"Error adding supplier: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': 'Failed to add supplier'}), 500


@app.route('/api/suppliers/<int:supplier_id>', methods=['DELETE'])
def delete_supplier(supplier_id):
    try:
        with get_db_context() as conn:
            conn.execute("UPDATE suppliers SET active=0 WHERE id=?", (supplier_id,))
            conn.commit()
        return jsonify({'status': 'ok'})
    except Exception as e:
        logger.error(f"Error deleting supplier {supplier_id}: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': 'Failed to delete supplier'}), 500


# ============================================================
# USER EDIT (update without delete) + report download alias
# ============================================================
@app.route('/api/users/<int:uid>', methods=['PUT'])
def update_user(uid):
    d = safe_json(request)
    try:
        with get_db_context() as conn:
            fields = []
            params = []
            for col in ('name', 'role', 'trade', 'department', 'create_scope'):
                if col in d:
                    fields.append(f"{col}=?")
                    params.append(d[col])
            for col in ('can_requisition', 'stores_access'):
                if col in d:
                    fields.append(f"{col}=?")
                    params.append(int(d[col]))
            if not fields:
                return jsonify({'status': 'error', 'message': 'No fields to update'}), 400
            params.append(uid)
            conn.execute(f"UPDATE users SET {', '.join(fields)} WHERE id=?", params)
            conn.commit()
        log_action('Admin', 'user_updated', f"User {uid} updated")
        return jsonify({'status': 'ok'})
    except Exception as e:
        logger.error(f"Error updating user {uid}: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': 'Failed to update user'}), 500


@app.route('/api/report/<rtype>/download')
def report_download_alias(rtype):
    """Alias for report downloads referenced by the frontend."""
    try:
        if rtype == 'weekly':
            return send_file(generate_weekly_report(), as_attachment=True)
        return jsonify({'status': 'error', 'message': f'Unknown report type: {rtype}'}), 404
    except Exception as e:
        logger.error(f"Error downloading report {rtype}: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': 'Failed to generate report'}), 500


@app.route('/api/report/weekly/preview')
def report_weekly_preview():
    """Lightweight preview data for the weekly report (in-page, no popup)."""
    try:
        week_start = _week_start_iso()
        cutoff = (date.today() - timedelta(days=7)).isoformat()
        with get_db_context() as conn:
            bd_count = conn.execute(
                "SELECT COUNT(*) FROM breakdowns WHERE log_date >= ?", (cutoff,)
            ).fetchone()[0]
            total_dt = conn.execute(
                "SELECT COALESCE(SUM(downtime_mins),0) FROM breakdowns WHERE log_date >= ?", (cutoff,)
            ).fetchone()[0]
            pm = conn.execute(
                "SELECT COUNT(*) total, SUM(CASE WHEN status='done' THEN 1 ELSE 0 END) done "
                "FROM pm_tasks WHERE week_start=?", (week_start,)
            ).fetchone()
        return jsonify({
            'breakdowns': bd_count,
            'total_downtime_hrs': round(total_dt / 60, 1),
            'pm_total': pm['total'] or 0,
            'pm_done': pm['done'] or 0,
        })
    except Exception as e:
        logger.error(f"Error building weekly preview: {e}", exc_info=True)
        return jsonify({}), 500



# ============================================================
# AREAS — plant area management (v10.2)
# ============================================================
@app.route('/api/areas')
def get_areas():
    """Get all active plant areas."""
    try:
        with get_db_context() as conn:
            rows = conn.execute(
                "SELECT * FROM areas WHERE active=1 ORDER BY sort_order, name"
            ).fetchall()
        return jsonify([dict(r) for r in rows])
    except Exception as e:
        logger.error(f"Error fetching areas: {e}", exc_info=True)
        return jsonify([]), 500


@app.route('/api/areas', methods=['POST'])
def add_area():
    d = safe_json(request)
    name = str(d.get('name', '')).strip()
    if not name:
        return jsonify({'status': 'error', 'message': 'Area name is required'}), 400
    try:
        with get_db_context() as conn:
            conn.execute(
                "INSERT INTO areas (name,code,description,sort_order) VALUES (?,?,?,?)",
                (name, d.get('code','').upper(), d.get('description',''),
                 int(d.get('sort_order', 99)))
            )
            conn.commit()
        return jsonify({'status': 'ok'})
    except Exception as e:
        logger.error(f"Error adding area: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': 'Failed to add area'}), 500


@app.route('/api/areas/<int:aid>', methods=['PUT'])
def update_area(aid):
    d = safe_json(request)
    try:
        with get_db_context() as conn:
            conn.execute(
                "UPDATE areas SET name=?, code=?, description=?, sort_order=? WHERE id=?",
                (d.get('name'), str(d.get('code','')).upper(),
                 d.get('description',''), d.get('sort_order', 99), aid)
            )
            conn.commit()
        return jsonify({'status': 'ok'})
    except Exception as e:
        logger.error(f"Error updating area {aid}: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': 'Failed to update area'}), 500


@app.route('/api/areas/<int:aid>/toggle', methods=['POST'])
def toggle_area(aid):
    try:
        with get_db_context() as conn:
            conn.execute(
                "UPDATE areas SET active = CASE WHEN active=1 THEN 0 ELSE 1 END WHERE id=?",
                (aid,)
            )
            conn.commit()
        return jsonify({'status': 'ok'})
    except Exception as e:
        logger.error(f"Error toggling area {aid}: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': 'Failed to toggle area'}), 500


# ============================================================
# EQUIPMENT CRUD (v10.2)
# ============================================================
@app.route('/api/equipment', methods=['POST'])
def add_equipment():
    """Admin adds a new machine."""
    d = safe_json(request)
    name = str(d.get('name', '')).strip()
    area = str(d.get('area', '')).strip()
    if not name or not area:
        return jsonify({'status': 'error', 'message': 'name and area are required'}), 400
    try:
        with get_db_context() as conn:
            cur = conn.execute("""
                INSERT INTO equipment
                (name, area, area_code, section, description, criticality,
                 planned_hrs_day, active)
                VALUES (?,?,?,?,?,?,?,1)
            """, (
                name, area,
                str(d.get('area_code', '')).upper(),
                d.get('section', area),
                d.get('description', ''),
                d.get('criticality', 'High'),
                float(d.get('planned_hrs_day', 18))
            ))
            eid = cur.lastrowid
            conn.commit()
        log_action('Admin', 'equipment_added', f"Equipment added: {name} ({area})")
        return jsonify({'status': 'ok', 'id': eid})
    except Exception as e:
        logger.error(f"Error adding equipment: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': 'Failed to add equipment'}), 500


@app.route('/api/equipment/<int:eid>', methods=['PUT'])
def update_equipment(eid):
    """Admin edits a machine."""
    d = safe_json(request)
    try:
        with get_db_context() as conn:
            fields, params = [], []
            for col in ('name', 'area', 'area_code', 'section', 'description',
                        'criticality', 'planned_hrs_day'):
                if col in d:
                    fields.append(f"{col}=?")
                    params.append(d[col])
            if not fields:
                return jsonify({'status': 'error', 'message': 'No fields to update'}), 400
            params.append(eid)
            conn.execute(f"UPDATE equipment SET {', '.join(fields)} WHERE id=?", params)
            conn.commit()
        log_action('Admin', 'equipment_updated', f"Equipment {eid} updated")
        return jsonify({'status': 'ok'})
    except Exception as e:
        logger.error(f"Error updating equipment {eid}: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': 'Failed to update equipment'}), 500


@app.route('/api/equipment/<int:eid>/deactivate', methods=['POST'])
def deactivate_equipment(eid):
    """Admin deactivates a machine (moves to inactive tab)."""
    try:
        with get_db_context() as conn:
            conn.execute("UPDATE equipment SET active=0 WHERE id=?", (eid,))
            conn.commit()
        log_action('Admin', 'equipment_deactivated', f"Equipment {eid} deactivated")
        return jsonify({'status': 'ok'})
    except Exception as e:
        logger.error(f"Error deactivating equipment {eid}: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': 'Failed to deactivate'}), 500


@app.route('/api/equipment/<int:eid>/reactivate', methods=['POST'])
def reactivate_equipment(eid):
    """Admin reactivates an inactive machine."""
    try:
        with get_db_context() as conn:
            conn.execute("UPDATE equipment SET active=1 WHERE id=?", (eid,))
            conn.commit()
        log_action('Admin', 'equipment_reactivated', f"Equipment {eid} reactivated")
        return jsonify({'status': 'ok'})
    except Exception as e:
        logger.error(f"Error reactivating equipment {eid}: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': 'Failed to reactivate'}), 500


@app.route('/api/equipment/inactive')
def get_inactive_equipment():
    """Get all inactive machines."""
    try:
        with get_db_context() as conn:
            rows = conn.execute(
                "SELECT * FROM equipment WHERE active=0 ORDER BY area, name"
            ).fetchall()
        return jsonify([dict(r) for r in rows])
    except Exception as e:
        logger.error(f"Error fetching inactive equipment: {e}", exc_info=True)
        return jsonify([]), 500


# ============================================================
# BREAKDOWN — NEW FLOW (v10.2)
# Supervisor logs (start = NOW) → artisan fixes freely →
# Supervisor marks complete (end = NOW) →
# Times pushed to artisan for review → agree or silent dispute
# ============================================================
@app.route('/api/breakdowns/<int:bd_id>/complete-supervisor', methods=['POST'])
def supervisor_complete_breakdown(bd_id):
    """Supervisor completes their side of a breakdown.
    If artisan hasn't filed yet → awaiting_artisan_review.
    If artisan already filed (artisan_completed) → compare times → complete or disputed."""
    d = safe_json(request)
    try:
        now = datetime.now().isoformat()
        end_hhmm = datetime.now().strftime('%H:%M')
        with get_db_context() as conn:
            row = conn.execute("SELECT * FROM breakdowns WHERE id=?", (bd_id,)).fetchone()
            if not row:
                return jsonify({'status': 'error', 'message': 'Not found'}), 404
            b = dict(row)
            art_done = b.get('status') == 'artisan_completed'

            disputed = 0
            dispute_field = None
            if art_done:
                art_start = (b.get('art_time_start') or '')[:5]
                art_end   = (b.get('art_time_end')   or '')[:5]
                sup_start = (b.get('sup_time_start') or '')[:5]
                if art_start and sup_start and art_start != sup_start:
                    disputed = 1; dispute_field = 'time_start'
                if art_end and end_hhmm and art_end != end_hhmm:
                    disputed = 1
                    dispute_field = (dispute_field + ', time_end') if dispute_field else 'time_end'

            new_status = ('disputed' if disputed else 'complete') if art_done else 'awaiting_artisan_review'

            # Recalculate downtime using supervisor's confirmed times
            b['art_time_start'] = b.get('art_time_start') or b.get('sup_time_start')
            b['art_time_end']   = end_hhmm
            downtime = _recalc_downtime(b) if art_done else None
            ftype = d.get('failure_type') or b.get('art_failure_type') or b.get('sup_failure_type')

            conn.execute("""
                UPDATE breakdowns
                SET status=?,
                    sup_completed_at=?,
                    sup_completed_by=?,
                    sup_time_end=?,
                    sup_failure_type=COALESCE(NULLIF(?,''), sup_failure_type),
                    sup_component=COALESCE(NULLIF(?,''), sup_component),
                    sup_failure_mode=COALESCE(NULLIF(?,''), sup_failure_mode),
                    sup_notes=COALESCE(NULLIF(?,''), sup_notes),
                    sup_machine_status=COALESCE(NULLIF(?,''), sup_machine_status),
                    disputed=?, dispute_field=?,
                    downtime_mins=COALESCE(?,downtime_mins),
                    production_downtime_mins=CASE WHEN ?='Production' THEN COALESCE(?,production_downtime_mins) ELSE production_downtime_mins END,
                    maintenance_downtime_mins=CASE WHEN ?!='Production' THEN COALESCE(?,maintenance_downtime_mins) ELSE maintenance_downtime_mins END
                WHERE id=? AND status IN ('supervisor_logged','artisan_completed')
            """, (new_status, now, d.get('supervisor_id'), end_hhmm,
                  d.get('failure_type'), d.get('component'), d.get('failure_mode'),
                  d.get('supervisor_notes'), d.get('machine_status'),
                  disputed, dispute_field,
                  downtime, ftype, downtime, ftype, downtime,
                  bd_id))
            conn.commit()
        log_action('Supervisor', 'breakdown_completed_sup',
                   f"Breakdown #{bd_id} marked resolved by supervisor")
        return jsonify({'status': 'ok'})
    except Exception as e:
        logger.error(f"Error completing breakdown {bd_id}: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': 'Failed to mark complete'}), 500


@app.route('/api/breakdowns/pending-review/<int:user_id>')
def get_breakdowns_pending_review(user_id):
    """Breakdowns awaiting artisan time review (supervisor already marked complete)."""
    try:
        with get_db_context() as conn:
            rows = conn.execute("""
                SELECT b.*, e.name as equip_name, e.area,
                       s.name as supervisor_name
                FROM breakdowns b
                LEFT JOIN equipment e ON e.id = b.equipment_id
                LEFT JOIN users s ON s.id = b.supervisor_id
                WHERE b.artisan_id=? AND b.status='awaiting_artisan_review'
                ORDER BY b.sup_completed_at DESC
            """, (user_id,)).fetchall()
        return jsonify([dict(r) for r in rows])
    except Exception as e:
        logger.error(f"Error fetching pending review breakdowns: {e}", exc_info=True)
        return jsonify([]), 500


@app.route('/api/breakdowns/<int:bd_id>/artisan-review', methods=['POST'])
def artisan_review_breakdown(bd_id):
    """Artisan reviews supervisor times. Agrees or disputes silently.
    Times are pre-filled from supervisor timestamps. If changed → silent dispute."""
    d = safe_json(request)
    try:
        with get_db_context() as conn:
            bd = conn.execute("SELECT * FROM breakdowns WHERE id=?", (bd_id,)).fetchone()
            if not bd:
                return jsonify({'status': 'error', 'message': 'Not found'}), 404
            bd = dict(bd)

            art_start = d.get('art_time_start', '')
            art_end   = d.get('art_time_end', '')

            # Use stored HH:MM fields — avoids datetime format ambiguity
            sup_start = (bd.get('sup_time_start') or '')[:5]
            # sup_time_end set by complete-supervisor; fall back to sup_completed_at
            sup_end = (bd.get('sup_time_end') or '')[:5]
            if not sup_end:
                ct = bd.get('sup_completed_at') or ''
                sup_end = ct[11:16] if 'T' in ct else ct[11:16]

            # Detect dispute: artisan changed either time
            disputed = 0
            if art_start and art_start != sup_start:
                disputed = 1
            if art_end and art_end != sup_end:
                disputed = 1

            # Recalculate downtime from final times
            def to_mins(t):
                try:
                    h, m = str(t).split(':')[:2]
                    return int(h)*60 + int(m)
                except Exception:
                    return None

            start_m = to_mins(art_start or sup_start)
            end_m   = to_mins(art_end or sup_end)
            downtime = None
            if start_m is not None and end_m is not None:
                downtime = end_m - start_m
                if downtime < 0:
                    downtime += 1440

            new_status = 'disputed' if disputed else 'complete'

            conn.execute("""
                UPDATE breakdowns SET
                    art_time_start=?, art_time_end=?,
                    art_failure_type=?, art_description=?, art_repair_action=?,
                    art_machine_status=?, art_followup=?,
                    art_failed_component=?, art_failure_mode=?,
                    art_completed_at=?, status=?,
                    disputed=?, downtime_mins=?,
                    production_downtime_mins=CASE WHEN ?='Production' THEN ? ELSE 0 END,
                    maintenance_downtime_mins=CASE WHEN ?='Production' THEN 0 ELSE ? END
                WHERE id=?
            """, (
                art_start, art_end,
                d.get('failure_type',''), d.get('description',''), d.get('repair_action',''),
                d.get('machine_status',''), d.get('followup',''),
                d.get('failed_component',''), d.get('failure_mode',''),
                datetime.now().isoformat(), new_status,
                disputed, downtime,
                bd.get('sup_failure_type'), downtime,
                bd.get('sup_failure_type'), downtime,
                bd_id
            ))
            # Parts used
            for p in (d.get('parts', []) or []):
                if p.get('desc'):
                    conn.execute(
                        "INSERT INTO breakdown_parts "
                        "(breakdown_id,part_desc,qty,logged_by,is_other,other_note) "
                        "VALUES (?,?,?,'artisan',?,?)",
                        (bd_id, p['desc'], p.get('qty',1),
                         1 if p.get('is_other') else 0,
                         p.get('other_note',''))
                    )
                    if not p.get('is_other'):
                        conn.execute(
                            "INSERT INTO reorder (part_desc,qty,source) VALUES (?,?,?)",
                            (p['desc'], p.get('qty',1), f'Breakdown #{bd_id}')
                        )
            conn.commit()

        log_action(d.get('artisan_name', 'Artisan'), 'breakdown_reviewed',
                   f"Breakdown #{bd_id} reviewed — {'DISPUTED' if disputed else 'agreed'}")
        return jsonify({'status': 'ok', 'disputed': bool(disputed)})
    except Exception as e:
        logger.error(f"Error in artisan review {bd_id}: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': 'Failed to submit review'}), 500


@app.route('/api/breakdowns/supervisor-active/<int:supervisor_id>')
def get_supervisor_active_breakdowns(supervisor_id):
    """Open breakdowns logged by this supervisor — showing Mark Complete button."""
    try:
        with get_db_context() as conn:
            rows = conn.execute("""
                SELECT b.*, e.name as equip_name, e.area,
                       a.name as artisan_name
                FROM breakdowns b
                LEFT JOIN equipment e ON e.id = b.equipment_id
                LEFT JOIN users a ON a.id = b.artisan_id
                WHERE b.supervisor_id=? AND b.status IN ('supervisor_logged','artisan_completed')
                ORDER BY b.logged_at DESC
            """, (supervisor_id,)).fetchall()
        return jsonify([dict(r) for r in rows])
    except Exception as e:
        logger.error(f"Error fetching supervisor active breakdowns: {e}", exc_info=True)
        return jsonify([]), 500


# ============================================================
# SPARES EXPORT — supplier-ready template (v10.2)
# ============================================================
@app.route('/api/equipment/<int:equip_id>/spares/export')
def export_spares_template(equip_id):
    """Export a blank spares-list Excel for sending to supplier to populate."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, PatternFill
        with get_db_context() as conn:
            equip = conn.execute(
                "SELECT * FROM equipment WHERE id=?", (equip_id,)
            ).fetchone()
            if not equip:
                return jsonify({'status': 'error', 'message': 'Equipment not found'}), 404
            existing = conn.execute(
                "SELECT * FROM equipment_spares WHERE equipment_id=? ORDER BY part_category, part_subcategory",
                (equip_id,)
            ).fetchall()

        wb = Workbook(); ws = wb.active
        ws.title = "Spares List"
        blue  = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
        green = PatternFill(start_color="166534", end_color="166534", fill_type="solid")
        wf    = Font(color="FFFFFF", bold=True)
        bold  = Font(bold=True)
        thin  = Side(style='thin', color='CCCCCC')
        bdr   = Border(left=thin, right=thin, top=thin, bottom=thin)

        ws.merge_cells('A1:H1')
        ws['A1'] = f"SPARE PARTS LIST — {equip['name']} ({equip['area']})"
        ws['A1'].font = Font(bold=True, size=14, color="1E3A5F")
        ws['A1'].alignment = Alignment(horizontal='center')

        ws['A3'] = "INSTRUCTIONS TO SUPPLIER:"; ws['A3'].font = Font(bold=True, color="CC0000")
        ws.merge_cells('B3:H3')
        ws['B3'] = ("Please complete ALL columns. Part numbers are mandatory. "
                    "Group by category. Return completed file for direct import.")
        ws['B3'].font = Font(italic=True, color="CC0000", size=10)

        headers = ['Category', 'Sub-Category', 'Part Description',
                   'Manufacturer Part No.', 'Std Qty', 'Unit', 'Supplier', 'Notes']
        hrow = 5
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=hrow, column=col, value=h)
            c.fill = blue; c.font = wf; c.border = bdr
            c.alignment = Alignment(horizontal='center')

        # Part categories as guide rows
        CATS = [
            ('Mechanical', 'Bearings'), ('Mechanical', 'Belts & Chains'),
            ('Mechanical', 'Gearboxes & Drives'), ('Mechanical', 'Seals & Gaskets'),
            ('Mechanical', 'Couplings & Shafts'), ('Mechanical', 'Wear Parts'),
            ('Electrical', 'Switchgear'), ('Electrical', 'Motors & Drives'),
            ('Electrical', 'Automation & Control'), ('Electrical', 'Sensors & Switches'),
            ('Hydraulic',  'Cylinders & Pumps'), ('Hydraulic', 'Hoses & Fittings'),
            ('Pneumatic',  'Cylinders & Valves'), ('Pneumatic', 'Air Treatment'),
            ('Consumables','Lubricants'), ('Consumables', 'Filters'),
        ]

        r = hrow + 1
        if existing:
            for item in existing:
                for col, val in enumerate([
                    item['part_category'], item['part_subcategory'],
                    item['part_desc'], item['part_number'] or '',
                    item['std_qty'], item['unit'], item['supplier'] or '', ''
                ], 1):
                    ws.cell(row=r, column=col, value=val).border = bdr
                r += 1
        else:
            # Blank template rows grouped by category
            for cat, sub in CATS:
                cat_fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
                for col, val in enumerate([cat, sub, '', '', '', 'each', '', ''], 1):
                    c = ws.cell(row=r, column=col, value=val)
                    c.border = bdr
                    if col <= 2: c.fill = cat_fill; c.font = Font(italic=True, color="666666")
                r += 1

        ws.column_dimensions['A'].width = 16
        ws.column_dimensions['B'].width = 22
        ws.column_dimensions['C'].width = 38
        ws.column_dimensions['D'].width = 22
        ws.column_dimensions['E'].width = 8
        ws.column_dimensions['F'].width = 8
        ws.column_dimensions['G'].width = 18
        ws.column_dimensions['H'].width = 20

        out_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'reports')
        os.makedirs(out_dir, exist_ok=True)
        safe_name = equip['name'].replace(' ', '_').replace('/', '-')
        fpath = os.path.join(out_dir, f"Spares_{safe_name}.xlsx")
        wb.save(fpath)
        return send_file(fpath, as_attachment=True,
                         download_name=f"Spares_{safe_name}.xlsx")
    except Exception as e:
        logger.error(f"Error exporting spares for {equip_id}: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': 'Failed to export spares'}), 500


@app.route('/api/equipment/<int:equip_id>/spares/import', methods=['POST'])
def import_spares(equip_id):
    """Import populated spares list from supplier Excel file."""
    try:
        from openpyxl import load_workbook
        import io
        file = request.files.get('file')
        if not file:
            return jsonify({'status': 'error', 'message': 'No file uploaded'}), 400
        wb = load_workbook(io.BytesIO(file.read()))
        ws = wb.active
        added = 0
        with get_db_context() as conn:
            for row in ws.iter_rows(min_row=6, values_only=True):
                cat, sub, desc, partno, qty, unit, supplier, notes = (list(row) + [None]*8)[:8]
                if not desc or str(desc).strip() == '':
                    continue
                conn.execute("""
                    INSERT INTO equipment_spares
                    (equipment_id, part_category, part_subcategory, part_desc,
                     part_number, std_qty, unit, supplier)
                    VALUES (?,?,?,?,?,?,?,?)
                """, (
                    equip_id,
                    str(cat or 'General').strip(),
                    str(sub or '').strip(),
                    str(desc).strip(),
                    str(partno or '').strip() or None,
                    float(qty or 1),
                    str(unit or 'each').strip(),
                    str(supplier or '').strip() or None,
                ))
                added += 1
            conn.commit()
        return jsonify({'status': 'ok', 'added': added})
    except Exception as e:
        logger.error(f"Error importing spares for {equip_id}: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': 'Import failed'}), 500


# ============================================================
# END v10.1 ADDED ROUTES
# ============================================================
# ============================================================
# REQUISITIONS (Purchase Requisitions) — NEW v10.1 ROLE-BASED
# ============================================================
def _next_pr_number(conn):
    year = date.today().year
    like = f"PR-{year}-%"
    row = conn.execute(
        "SELECT pr_number FROM requisitions WHERE pr_number LIKE ? ORDER BY id DESC LIMIT 1",
        (like,)
    ).fetchone()
    seq = 1
    if row and row['pr_number']:
        try:
            seq = int(row['pr_number'].split('-')[-1]) + 1
        except (ValueError, IndexError):
            seq = 1
    return f"PR-{year}-{seq:06d}"


@app.route('/api/requisitions')
def get_requisitions():
    """Get requisitions based on user role."""
    try:
        user_role = request.args.get('user_role', 'manager')
        with get_db_context() as conn:
            if user_role in ('stores_admin', 'admin'):
                rows = conn.execute("""
                    SELECT r.*, e.name as equip_name, u.name as raised_by_name,
                           (SELECT COUNT(*) FROM requisition_items WHERE requisition_id=r.id) as item_count
                    FROM requisitions r
                    LEFT JOIN equipment e ON e.id = r.equipment_id
                    LEFT JOIN users u ON u.id = r.raised_by
                    ORDER BY r.submitted_at DESC
                """).fetchall()
            else:
                rows = conn.execute("""
                    SELECT r.*, e.name as equip_name, u.name as raised_by_name,
                           (SELECT COUNT(*) FROM requisition_items WHERE requisition_id=r.id) as item_count
                    FROM requisitions r
                    LEFT JOIN equipment e ON e.id = r.equipment_id
                    LEFT JOIN users u ON u.id = r.raised_by
                    WHERE r.raised_by = (SELECT id FROM users WHERE name=? LIMIT 1) OR r.status != 'submitted'
                    ORDER BY r.submitted_at DESC
                """, (request.args.get('user_name', ''),)).fetchall()
        return jsonify([dict(r) for r in rows])
    except Exception as e:
        logger.error(f"Error fetching requisitions: {e}", exc_info=True)
        return jsonify([]), 500


@app.route('/api/requisitions', methods=['POST'])
def add_requisition():
    """Create requisition. Only admin/manager/supervisor can create."""
    d = safe_json(request)
    user_role = d.get('user_role')
    if user_role not in ('admin', 'manager', 'supervisor', 'team_leader'):
        return jsonify({'status': 'error', 'message': 'Only managers, team leaders and supervisors can create requisitions'}), 403
    items = d.get('items', [])
    if not items:
        return jsonify({'status': 'error', 'message': 'At least one line item is required'}), 400
    for it in items:
        if not it.get('part_desc'):
            return jsonify({'status': 'error', 'message': 'All items must have a description'}), 400
        if not it.get('qty'):
            return jsonify({'status': 'error', 'message': 'All items must have a quantity'}), 400
        if not it.get('planned_install_date'):
            return jsonify({'status': 'error', 'message': 'All items must have a planned installation date'}), 400
    try:
        with get_db_context() as conn:
            pr_number = _next_pr_number(conn)
            cur = conn.execute("""
                INSERT INTO requisitions
                (pr_number, raised_by, raised_by_name, department, equipment_id,
                 motivation, estimated_value, is_special_order, status)
                VALUES (?,?,?,?,?,?,?,?,'submitted')
            """, (
                pr_number, d.get('raised_by'), d.get('raised_by_name'),
                d.get('department'), d.get('equipment_id'), d.get('motivation'),
                d.get('estimated_value'), int(d.get('is_special_order', 0) or 0)
            ))
            pr_id = cur.lastrowid
            for it in items:
                conn.execute("""
                    INSERT INTO requisition_items
                    (requisition_id, part_number, part_desc, qty, planned_install_date, unit, is_special, stock_id)
                    VALUES (?,?,?,?,?,?,?,?)
                """, (
                    pr_id, it.get('part_number'), it.get('part_desc'),
                    it.get('qty', 1), it.get('planned_install_date'),
                    it.get('unit', 'each'),
                    int(it.get('is_special', 0) or 0), it.get('stock_id')
                ))
            conn.commit()
        log_action(d.get('raised_by_name', 'Manager'), 'requisition_submitted', pr_number)
        return jsonify({'status': 'ok', 'pr_number': pr_number, 'id': pr_id})
    except Exception as e:
        logger.error(f"Error adding requisition: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': 'Failed to submit requisition'}), 500


@app.route('/api/requisitions/<int:pr_id>')
def get_requisition_detail(pr_id):
    try:
        with get_db_context() as conn:
            pr = conn.execute("""
                SELECT r.*, e.name as equip_name, u.name as raised_by_name
                FROM requisitions r
                LEFT JOIN equipment e ON e.id = r.equipment_id
                LEFT JOIN users u ON u.id = r.raised_by
                WHERE r.id=?
            """, (pr_id,)).fetchone()
            if not pr:
                return jsonify({'status': 'error', 'message': 'Not found'}), 404
            items = conn.execute(
                "SELECT * FROM requisition_items WHERE requisition_id=?", (pr_id,)
            ).fetchall()
            status_updates = conn.execute("""
                SELECT status_update, updated_by_name, updated_at
                FROM requisition_status
                WHERE requisition_id=?
                ORDER BY updated_at DESC
            """, (pr_id,)).fetchall()
        result = dict(pr)
        result['items'] = [dict(i) for i in items]
        result['status_history'] = [dict(s) for s in status_updates]
        return jsonify(result)
    except Exception as e:
        logger.error(f"Error fetching requisition {pr_id}: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': 'Failed to fetch requisition'}), 500


@app.route('/api/requisitions/<int:pr_id>/action', methods=['POST'])
def action_requisition(pr_id):
    """ONLY stores_admin (Steven) can action a requisition."""
    d = safe_json(request)
    if d.get('user_role') != 'stores_admin':
        return jsonify({'status': 'error', 'message': 'Only the Stores Admin can action requisitions'}), 403
    new_status = d.get('new_status', 'ordered')
    valid_statuses = ['pending_approval', 'pending_quotation', 'ordered', 'declined', 'delivered', 'partially_received']
    if new_status not in valid_statuses:
        return jsonify({'status': 'error', 'message': f'Invalid status: {new_status}'}), 400
    try:
        with get_db_context() as conn:
            extra_fields = ""
            extra_vals = []
            # Auto-generate quote_ref when marking as ordered
            if new_status == 'ordered':
                quote_ref = f"QT-{datetime.now().year}-{str(pr_id).zfill(6)}"
                supplier  = d.get('supplier_name', '').strip() or None
                extra_fields = ", quote_ref=?, supplier_name=?"
                extra_vals   = [quote_ref, supplier]
            conn.execute(f"""
                UPDATE requisitions
                SET status=?, actioned_by=?, actioned_at=?, action_notes=?{extra_fields}
                WHERE id=?
            """, (new_status, d.get('actioned_by'), datetime.now().isoformat(),
                  d.get('notes'), *extra_vals, pr_id))
            conn.commit()
            if new_status == 'ordered':
                row = conn.execute("SELECT quote_ref FROM requisitions WHERE id=?", (pr_id,)).fetchone()
                quote_ref = row['quote_ref'] if row else None
            else:
                quote_ref = None
        log_action('Stores Admin', 'requisition_actioned', f"PR #{pr_id} → {new_status}")
        return jsonify({'status': 'ok', 'quote_ref': quote_ref})
    except Exception as e:
        logger.error(f"Error actioning requisition {pr_id}: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': 'Failed to action requisition'}), 500


@app.route('/api/requisitions/search-by-ref')
def search_requisition_by_ref():
    """Search open requisitions by quote_ref, invoice_ref or pr_number. Used by storeman receive screen."""
    q = (request.args.get('q') or '').strip()
    if not q:
        return jsonify([])
    try:
        with get_db_context() as conn:
            rows = conn.execute("""
                SELECT r.*, e.name as equip_name, u.name as raised_by_name,
                       (SELECT COUNT(*) FROM requisition_items WHERE requisition_id=r.id) as item_count
                FROM requisitions r
                LEFT JOIN equipment e ON e.id = r.equipment_id
                LEFT JOIN users u ON u.id = r.raised_by
                WHERE (r.quote_ref LIKE ? OR r.invoice_ref LIKE ? OR r.pr_number LIKE ?)
                  AND r.status IN ('ordered','partially_received')
                ORDER BY r.submitted_at DESC LIMIT 10
            """, (f'%{q}%', f'%{q}%', f'%{q}%')).fetchall()
        return jsonify([dict(r) for r in rows])
    except Exception as e:
        logger.error(f"Error searching requisitions: {e}", exc_info=True)
        return jsonify([]), 500


@app.route('/api/requisitions/<int:pr_id>/receive', methods=['POST'])
def receive_requisition(pr_id):
    """Storeman records delivery against a requisition. Updates status and records invoice_ref."""
    d = safe_json(request)
    invoice_ref = (d.get('invoice_ref') or '').strip() or None
    partial     = bool(d.get('partial', False))
    notes       = d.get('notes') or None
    try:
        with get_db_context() as conn:
            pr = conn.execute("SELECT * FROM requisitions WHERE id=?", (pr_id,)).fetchone()
            if not pr:
                return jsonify({'status': 'error', 'message': 'Requisition not found'}), 404
            new_status = 'partially_received' if partial else 'delivered'
            conn.execute("""
                UPDATE requisitions
                SET status=?, invoice_ref=?, actioned_at=?, action_notes=?
                WHERE id=?
            """, (new_status, invoice_ref, datetime.now().isoformat(), notes, pr_id))
            # Record a stock movement for each item that has a stock_id linked
            items = conn.execute("""
                SELECT * FROM requisition_items WHERE requisition_id=?
            """, (pr_id,)).fetchall()
            for item in items:
                if item['stock_id']:
                    stock_row = conn.execute(
                        "SELECT qty_on_hand FROM stores_stock WHERE id=?", (item['stock_id'],)
                    ).fetchone()
                    if stock_row:
                        qty_recv = float(d.get(f'qty_{item["id"]}') or item['qty'])
                        new_qty  = stock_row['qty_on_hand'] + qty_recv
                        conn.execute("UPDATE stores_stock SET qty_on_hand=? WHERE id=?",
                                     (new_qty, item['stock_id']))
                        conn.execute("""
                            INSERT INTO stock_movements
                            (stock_id, movement_type, qty, qty_after, reference, notes,
                             logged_by, logged_by_name, requisition_id)
                            VALUES (?,?,?,?,?,?,?,?,?)
                        """, (item['stock_id'], 'receipt', qty_recv, new_qty,
                              invoice_ref or pr['quote_ref'], notes,
                              d.get('received_by'), d.get('received_by_name', 'Storeman'), pr_id))
            conn.commit()
        log_action(d.get('received_by_name','Storeman'), 'stock_received',
                   f"PR {pr['pr_number']} → {new_status} (inv: {invoice_ref})")
        return jsonify({'status': 'ok', 'new_status': new_status})
    except Exception as e:
        logger.error(f"Error receiving requisition {pr_id}: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': 'Failed to record receipt'}), 500


@app.route('/api/requisitions/<int:pr_id>/request-status', methods=['POST'])
def request_requisition_status(pr_id):
    """Admin or stores_admin can query/flag a requisition."""
    d = safe_json(request)
    if d.get('user_role') not in ('admin', 'stores_admin'):
        return jsonify({'status': 'error', 'message': 'Only Admin or Stores Admin can query requisitions'}), 403
    try:
        with get_db_context() as conn:
            conn.execute("UPDATE requisitions SET admin_queried=1 WHERE id=?", (pr_id,))
            conn.commit()
        actor = 'Stores Admin' if d.get('user_role') == 'stores_admin' else 'Admin'
        log_action(actor, 'requisition_queried', f"PR #{pr_id} flagged as queried")
        return jsonify({'status': 'ok', 'message': 'Requisition flagged as queried'})
    except Exception as e:
        logger.error(f"Error querying requisition {pr_id}: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': 'Failed to flag requisition'}), 500


@app.route('/api/requisitions/<int:pr_id>/status', methods=['POST'])
def update_requisition_status(pr_id):
    """Stores Admin updates status."""
    d = safe_json(request)
    if d.get('user_role') != 'stores_admin':
        return jsonify({'status': 'error', 'message': 'Only Stores Admin can post status updates'}), 403
    status_update = d.get('status_update')
    valid_statuses = ['Awaiting Pricing', 'In For Approval', 'Awaiting Delivery']
    if status_update not in valid_statuses:
        return jsonify({'status': 'error', 'message': f'Invalid status. Must be one of: {", ".join(valid_statuses)}'}), 400
    try:
        with get_db_context() as conn:
            conn.execute("""
                INSERT INTO requisition_status
                (requisition_id, status_update, updated_by, updated_by_name, updated_at)
                VALUES (?,?,?,?,?)
            """, (pr_id, status_update, d.get('updated_by'), d.get('updated_by_name', 'Steven'),
                  datetime.now().isoformat()))
            conn.commit()
        log_action('Stores Admin', 'status_updated', f"PR #{pr_id} status: {status_update}")
        return jsonify({'status': 'ok', 'message': f'Status updated to: {status_update}'})
    except Exception as e:
        logger.error(f"Error updating requisition status {pr_id}: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': 'Failed to update status'}), 500


@app.route('/api/requisitions/<int:pr_id>/status-history')
def get_requisition_status_history(pr_id):
    try:
        with get_db_context() as conn:
            updates = conn.execute("""
                SELECT status_update, updated_by_name, updated_at
                FROM requisition_status
                WHERE requisition_id=?
                ORDER BY updated_at DESC
            """, (pr_id,)).fetchall()
        return jsonify([dict(u) for u in updates])
    except Exception as e:
        logger.error(f"Error fetching status history for {pr_id}: {e}", exc_info=True)
        return jsonify([]), 500


@app.route('/api/requisitions/<int:pr_id>/rfq')
def requisition_rfq(pr_id):
    """Export a requisition as an RFQ Excel file."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        with get_db_context() as conn:
            pr = conn.execute("""
                SELECT r.*, e.name as equip_name, u.name as raised_by_name
                FROM requisitions r
                LEFT JOIN equipment e ON e.id = r.equipment_id
                LEFT JOIN users u ON u.id = r.raised_by
                WHERE r.id=?
            """, (pr_id,)).fetchone()
            if not pr:
                return jsonify({'status': 'error', 'message': 'Not found'}), 404
            items = conn.execute(
                "SELECT * FROM requisition_items WHERE requisition_id=?", (pr_id,)
            ).fetchall()
        wb = Workbook(); ws = wb.active; ws.title = "RFQ"
        blue = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        white_bold = Font(color="FFFFFF", bold=True, size=12)
        bold = Font(bold=True)
        thin = Side(style='thin', color='CCCCCC')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        ws.merge_cells('A1:F1')
        ws['A1'] = "REQUEST FOR QUOTATION"
        ws['A1'].font = Font(bold=True, size=16, color="2563EB")
        ws['A1'].alignment = Alignment(horizontal='center')
        ws['A3'] = "PR Number:"; ws['A3'].font = bold; ws['B3'] = pr['pr_number']
        ws['A4'] = "Raised By:"; ws['A4'].font = bold; ws['B4'] = pr['raised_by_name'] or ''
        ws['A5'] = "Department:"; ws['A5'].font = bold; ws['B5'] = pr['department'] or ''
        ws['A6'] = "Machine:"; ws['A6'].font = bold; ws['B6'] = pr['equip_name'] or 'N/A'
        ws['A7'] = "Date:"; ws['A7'].font = bold; ws['B7'] = (pr['submitted_at'] or '')[:10]
        ws.merge_cells('A9:F9')
        ws['A9'] = ("IMPORTANT: Manufacturer part numbers are CRITICAL and MUST be included in your quote.")
        ws['A9'].font = Font(bold=True, color="EF4444", size=10)
        ws['A9'].alignment = Alignment(wrap_text=True, vertical='top')
        ws.row_dimensions[9].height = 45
        headers = ['#', 'Part Description', 'Part Number', 'Qty', 'Unit', 'Install Date']
        hrow = 11
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=hrow, column=col, value=h)
            cell.fill = blue; cell.font = white_bold; cell.border = border
            cell.alignment = Alignment(horizontal='center')
        r = hrow + 1
        for i, item in enumerate(items, 1):
            ws.cell(row=r, column=1, value=i).border = border
            ws.cell(row=r, column=2, value=item['part_desc']).border = border
            ws.cell(row=r, column=3, value=item['part_number'] or '').border = border
            ws.cell(row=r, column=4, value=item['qty']).border = border
            ws.cell(row=r, column=5, value=item['unit'] or 'each').border = border
            ws.cell(row=r, column=6, value=(item['planned_install_date'] or '')[:10]).border = border
            r += 1
        ws.cell(row=hrow, column=7, value='Unit Price (quote)').fill = blue
        ws.cell(row=hrow, column=7).font = white_bold; ws.cell(row=hrow, column=7).border = border
        ws.column_dimensions['A'].width = 5; ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 22; ws.column_dimensions['D'].width = 8
        ws.column_dimensions['E'].width = 10; ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 18
        if pr['motivation']:
            mrow = r + 1
            ws.cell(row=mrow, column=1, value="Motivation:").font = bold
            ws.merge_cells(start_row=mrow, start_column=2, end_row=mrow, end_column=7)
            ws.cell(row=mrow, column=2, value=pr['motivation'])
        out_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'reports')
        os.makedirs(out_dir, exist_ok=True)
        fpath = os.path.join(out_dir, f"RFQ_{pr['pr_number']}.xlsx")
        wb.save(fpath)
        return send_file(fpath, as_attachment=True, download_name=f"RFQ_{pr['pr_number']}.xlsx")
    except Exception as e:
        logger.error(f"Error generating RFQ for {pr_id}: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': 'Failed to generate RFQ'}), 500




# ============================================================
# COMPONENT TRACKING
# ============================================================

@app.route('/api/component-categories')
def get_component_categories():
    try:
        with get_db_context() as conn:
            rows = conn.execute(
                "SELECT * FROM component_categories ORDER BY name"
            ).fetchall()
        return jsonify([dict(r) for r in rows])
    except Exception as e:
        logger.error(f"Error fetching component categories: {e}", exc_info=True)
        return jsonify([]), 500


@app.route('/api/component-categories', methods=['POST'])
def add_component_category():
    d = safe_json(request)
    name = str(d.get('name', '')).strip()
    if not name:
        return jsonify({'status': 'error', 'message': 'name required'}), 400
    try:
        with get_db_context() as conn:
            conn.execute(
                "INSERT INTO component_categories (name, description, is_wear_item, default_lifetime_days) VALUES (?,?,?,?)",
                (name, d.get('description'), int(d.get('is_wear_item', 1) or 1), d.get('default_lifetime_days'))
            )
            conn.commit()
        return jsonify({'status': 'ok'})
    except Exception as e:
        logger.error(f"Error adding component category: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': 'Failed to add category'}), 500


@app.route('/api/component-categories/<int:cid>', methods=['DELETE'])
def delete_component_category(cid):
    try:
        with get_db_context() as conn:
            conn.execute("DELETE FROM component_categories WHERE id=?", (cid,))
            conn.commit()
        return jsonify({'status': 'ok'})
    except Exception as e:
        logger.error(f"Error deleting component category {cid}: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': 'Failed to delete category'}), 500


@app.route('/api/equipment/<int:eid>/components')
def get_equipment_components(eid):
    try:
        with get_db_context() as conn:
            rows = conn.execute("""
                SELECT ec.*, cc.name as category_name, cc.is_wear_item as cat_is_wear
                FROM equipment_components ec
                LEFT JOIN component_categories cc ON cc.id = ec.category_id
                WHERE ec.equipment_id = ? AND ec.active = 1
                ORDER BY ec.name
            """, (eid,)).fetchall()
        return jsonify([dict(r) for r in rows])
    except Exception as e:
        logger.error(f"Error fetching components for equipment {eid}: {e}", exc_info=True)
        return jsonify([]), 500


@app.route('/api/equipment/<int:eid>/components', methods=['POST'])
def add_equipment_component(eid):
    d = safe_json(request)
    name = str(d.get('name', '')).strip()
    if not name:
        return jsonify({'status': 'error', 'message': 'Component name required'}), 400
    try:
        with get_db_context() as conn:
            conn.execute("""
                INSERT INTO equipment_components
                (equipment_id, category_id, name, location_on_machine, part_number,
                 is_wear_item, estimated_lifetime_days, condition, created_by)
                VALUES (?,?,?,?,?,?,?,?,?)
            """, (
                eid, d.get('category_id'), name, d.get('location_on_machine'),
                d.get('part_number'), int(d.get('is_wear_item', 1) or 1),
                d.get('estimated_lifetime_days'),
                d.get('condition', 'Unknown'), d.get('created_by')
            ))
            conn.commit()
        return jsonify({'status': 'ok'})
    except Exception as e:
        logger.error(f"Error adding component: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': 'Failed to add component'}), 500


@app.route('/api/components/<int:cid>', methods=['DELETE'])
def delete_component(cid):
    try:
        with get_db_context() as conn:
            conn.execute("UPDATE equipment_components SET active=0 WHERE id=?", (cid,))
            conn.commit()
        return jsonify({'status': 'ok'})
    except Exception as e:
        logger.error(f"Error deleting component {cid}: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': 'Failed to delete component'}), 500


@app.route('/api/components/<int:cid>/assess', methods=['POST'])
def assess_component(cid):
    """Record or update a condition assessment. Valid only before first verified replacement."""
    d = safe_json(request)
    condition = d.get('condition', 'Unknown')
    if condition not in ('Unknown', 'New', 'Good', 'Fair', 'Poor', 'Critical'):
        return jsonify({'status': 'error', 'message': 'Invalid condition value'}), 400
    try:
        with get_db_context() as conn:
            conn.execute("""
                UPDATE equipment_components SET
                    condition             = ?,
                    estimated_age_days    = ?,
                    estimated_remaining_days = ?,
                    confidence_level      = ?,
                    condition_notes       = ?,
                    assessed_by           = ?,
                    assessed_by_name      = ?,
                    assessment_date       = ?
                WHERE id = ?
            """, (
                condition,
                d.get('estimated_age_days'),
                d.get('estimated_remaining_days'),
                d.get('confidence_level'),
                d.get('notes'),
                d.get('assessed_by'),
                d.get('assessed_by_name'),
                d.get('assessment_date') or datetime.now().strftime('%Y-%m-%d'),
                cid
            ))
            conn.execute("""
                INSERT INTO component_events
                (component_id, event_type, event_date, notes, logged_by, logged_by_name)
                VALUES (?,?,?,?,?,?)
            """, (
                cid, 'assessment',
                d.get('assessment_date') or datetime.now().strftime('%Y-%m-%d'),
                f"Condition: {condition}. {d.get('notes') or ''}".strip('. '),
                d.get('assessed_by'), d.get('assessed_by_name')
            ))
            conn.commit()
        return jsonify({'status': 'ok'})
    except Exception as e:
        logger.error(f"Error assessing component {cid}: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': 'Failed to save assessment'}), 500


@app.route('/api/components/<int:cid>/replace', methods=['POST'])
def log_component_replacement(cid):
    """Log a verified replacement — resets the heartbeat and becomes authoritative source."""
    d = safe_json(request)
    event_date = d.get('event_date') or datetime.now().strftime('%Y-%m-%d')
    try:
        with get_db_context() as conn:
            row = conn.execute(
                "SELECT replacement_count, failure_count FROM equipment_components WHERE id=?", (cid,)
            ).fetchone()
            if not row:
                return jsonify({'status': 'error', 'message': 'Component not found'}), 404
            new_count = (row['replacement_count'] or 0) + 1
            conn.execute("""
                UPDATE equipment_components SET
                    condition            = 'New',
                    installation_date    = ?,
                    replacement_count    = ?,
                    last_replacement     = ?,
                    estimated_age_days   = 0,
                    estimated_remaining_days = estimated_lifetime_days,
                    condition_notes      = NULL,
                    assessed_by          = NULL,
                    assessed_by_name     = NULL,
                    assessment_date      = NULL
                WHERE id = ?
            """, (event_date, new_count, event_date, cid))
            conn.execute("""
                INSERT INTO component_events
                (component_id, event_type, event_date, work_order_ref, part_number, supplier, notes, logged_by, logged_by_name)
                VALUES (?,?,?,?,?,?,?,?,?)
            """, (
                cid, 'replacement', event_date,
                d.get('work_order_ref'), d.get('part_number'), d.get('supplier'),
                d.get('notes'), d.get('logged_by'), d.get('logged_by_name')
            ))
            conn.commit()
        return jsonify({'status': 'ok'})
    except Exception as e:
        logger.error(f"Error logging replacement for component {cid}: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': 'Failed to log replacement'}), 500


@app.route('/api/components/<int:cid>/failure', methods=['POST'])
def log_component_failure(cid):
    """Log a component failure event."""
    d = safe_json(request)
    event_date = d.get('event_date') or datetime.now().strftime('%Y-%m-%d')
    try:
        with get_db_context() as conn:
            row = conn.execute(
                "SELECT failure_count FROM equipment_components WHERE id=?", (cid,)
            ).fetchone()
            if not row:
                return jsonify({'status': 'error', 'message': 'Component not found'}), 404
            new_count = (row['failure_count'] or 0) + 1
            conn.execute("""
                UPDATE equipment_components SET
                    condition      = 'Critical',
                    failure_count  = ?,
                    last_failure   = ?
                WHERE id = ?
            """, (new_count, event_date, cid))
            conn.execute("""
                INSERT INTO component_events
                (component_id, event_type, event_date, work_order_ref, notes, logged_by, logged_by_name)
                VALUES (?,?,?,?,?,?,?)
            """, (
                cid, 'failure', event_date,
                d.get('work_order_ref'), d.get('notes'),
                d.get('logged_by'), d.get('logged_by_name')
            ))
            conn.commit()
        return jsonify({'status': 'ok'})
    except Exception as e:
        logger.error(f"Error logging failure for component {cid}: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': 'Failed to log failure'}), 500


@app.route('/api/components/<int:cid>/history')
def get_component_history(cid):
    try:
        with get_db_context() as conn:
            rows = conn.execute("""
                SELECT * FROM component_events
                WHERE component_id = ?
                ORDER BY logged_at DESC
            """, (cid,)).fetchall()
        return jsonify([dict(r) for r in rows])
    except Exception as e:
        logger.error(f"Error fetching component history {cid}: {e}", exc_info=True)
        return jsonify([]), 500


@app.route('/api/components/all')
def get_all_components():
    """Return all active components across all equipment, for dashboards."""
    try:
        with get_db_context() as conn:
            rows = conn.execute("""
                SELECT ec.*, cc.name as category_name, e.name as equipment_name, e.area
                FROM equipment_components ec
                LEFT JOIN component_categories cc ON cc.id = ec.category_id
                LEFT JOIN equipment e ON e.id = ec.equipment_id
                WHERE ec.active = 1
                ORDER BY e.name, ec.name
            """).fetchall()
        return jsonify([dict(r) for r in rows])
    except Exception as e:
        logger.error(f"Error fetching all components: {e}", exc_info=True)
        return jsonify([]), 500


# ---- Plant Hierarchy (Phase 2) ----

def _seed_plant_hierarchy(conn):
    """Seed dept→machine tree from equipment.dept when plant_nodes is empty.
    Uses the human-readable dept field (Wetmill, Drymill, Kilns, etc.).
    Called automatically the first time the hierarchy API is queried.
    """
    depts = conn.execute(
        "SELECT DISTINCT dept FROM equipment WHERE active=1 ORDER BY dept"
    ).fetchall()
    for row in depts:
        dept_name = row['dept'] or 'General'
        conn.execute(
            "INSERT INTO plant_nodes (parent_id, name, label) VALUES (NULL,?,?)",
            (dept_name, dept_name)
        )
        dept_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        conn.execute(
            "INSERT OR IGNORE INTO plant_node_ancestors "
            "(ancestor_id, descendant_id, depth) VALUES (?,?,0)",
            (dept_id, dept_id)
        )
        machines = conn.execute(
            "SELECT id, name, compound_id, equipment_id FROM equipment "
            "WHERE dept=? AND active=1 ORDER BY name",
            (dept_name,)
        ).fetchall()
        for m in machines:
            cid = m['compound_id'] or m['equipment_id'] or ''
            conn.execute(
                "INSERT INTO plant_nodes (parent_id, name, equipment_id) VALUES (?,?,?)",
                (dept_id, m['name'], cid)
            )
            leaf_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
            conn.execute(
                "INSERT OR IGNORE INTO plant_node_ancestors "
                "(ancestor_id, descendant_id, depth) VALUES (?,?,0)",
                (leaf_id, leaf_id)
            )
            conn.execute("""
                INSERT INTO plant_node_ancestors (ancestor_id, descendant_id, depth)
                SELECT ancestor_id, ?, depth + 1
                FROM plant_node_ancestors WHERE descendant_id = ?
            """, (leaf_id, dept_id))
            conn.execute(
                "UPDATE equipment SET node_id=? WHERE id=?",
                (leaf_id, m['id'])
            )
    conn.commit()


@app.route('/api/plant/nodes')
def api_plant_roots():
    """Return top-level plant nodes. Auto-seeds from equipment areas on first call."""
    try:
        with get_db_context() as conn:
            count = conn.execute("SELECT COUNT(*) FROM plant_nodes").fetchone()[0]
            if count == 0:
                _seed_plant_hierarchy(conn)
            rows = conn.execute("""
                SELECT n.*,
                       (SELECT COUNT(*) FROM plant_nodes c
                        WHERE c.parent_id = n.id AND c.active=1) AS child_count
                FROM plant_nodes n
                WHERE n.parent_id IS NULL AND n.active=1
                ORDER BY n.sort_order, n.name
            """).fetchall()
            return jsonify([dict(r) for r in rows])
    except Exception as e:
        logger.error(f"api_plant_roots error: {e}", exc_info=True)
        return jsonify([]), 500


@app.route('/api/plant/nodes/<int:node_id>/children')
def api_plant_children(node_id):
    """Return direct children of a plant node, each with child_count and equipment info."""
    try:
        with get_db_context() as conn:
            rows = conn.execute("""
                SELECT n.*,
                       (SELECT COUNT(*) FROM plant_nodes c
                        WHERE c.parent_id = n.id AND c.active=1) AS child_count,
                       e.id AS db_equip_id, e.area, e.criticality, e.planned_hrs_day
                FROM plant_nodes n
                LEFT JOIN equipment e ON e.equipment_id = n.equipment_id AND e.active=1
                WHERE n.parent_id = ? AND n.active=1
                ORDER BY n.sort_order, n.name
            """, (node_id,)).fetchall()
            return jsonify([dict(r) for r in rows])
    except Exception as e:
        logger.error(f"api_plant_children error: {e}", exc_info=True)
        return jsonify([]), 500


@app.route('/api/plant/nodes/<int:node_id>/breadcrumb')
def api_plant_breadcrumb(node_id):
    """Return the full path from root to node_id, ordered root-first."""
    try:
        from database import get_node_breadcrumb
        return jsonify(get_node_breadcrumb(node_id))
    except Exception as e:
        logger.error(f"api_plant_breadcrumb error: {e}", exc_info=True)
        return jsonify([]), 500


@app.route('/api/plant/nodes', methods=['POST'])
def api_plant_add_node():
    """Add a new plant node. Admin only."""
    if not session.get('user') or session['user'].get('role') != 'admin':
        return jsonify({'error': 'Admin only'}), 403
    try:
        d = request.get_json() or {}
        name = (d.get('name') or '').strip()
        if not name:
            return jsonify({'error': 'name required'}), 400
        from database import add_plant_node
        new_id = add_plant_node(
            parent_id=d.get('parent_id'),
            name=name,
            label=d.get('label'),
            equipment_id=d.get('equipment_id') or None,
            notes=d.get('notes')
        )
        return jsonify({'id': new_id})
    except Exception as e:
        logger.error(f"api_plant_add_node error: {e}", exc_info=True)
        return jsonify({'error': 'Failed to add node'}), 500


@app.route('/api/plant/nodes/<int:node_id>/move', methods=['PUT'])
def api_plant_move_node(node_id):
    """Move a plant node (and its subtree) to a new parent. Admin only."""
    if not session.get('user') or session['user'].get('role') != 'admin':
        return jsonify({'error': 'Admin only'}), 403
    try:
        d = request.get_json() or {}
        new_parent_id = d.get('parent_id')  # None = move to root
        from database import move_node
        ok = move_node(node_id, new_parent_id)
        return jsonify({'ok': ok})
    except Exception as e:
        logger.error(f"api_plant_move_node error: {e}", exc_info=True)
        return jsonify({'error': 'Failed to move node'}), 500


@app.route('/api/plant/by-equipment/<int:equip_db_id>')
def api_plant_by_equipment(equip_db_id):
    """Return the plant_node that is currently linked to an equipment row, if any."""
    try:
        with get_db_context() as conn:
            row = conn.execute(
                "SELECT pn.* FROM plant_nodes pn "
                "JOIN equipment e ON e.node_id = pn.id "
                "WHERE e.id = ? AND pn.active = 1",
                (equip_db_id,)
            ).fetchone()
            return jsonify(dict(row) if row else {})
    except Exception as e:
        logger.error(f"api_plant_by_equipment error: {e}", exc_info=True)
        return jsonify({}), 500


@app.route('/api/plant/nodes/<int:node_id>', methods=['PATCH'])
def api_plant_patch_node(node_id):
    """Rename a plant node or update its label. Admin only."""
    if not session.get('user') or session['user'].get('role') != 'admin':
        return jsonify({'error': 'Admin only'}), 403
    try:
        d = request.get_json() or {}
        with get_db_context() as conn:
            if 'name' in d:
                conn.execute(
                    "UPDATE plant_nodes SET name=? WHERE id=?",
                    (d['name'].strip(), node_id)
                )
            if 'label' in d:
                conn.execute(
                    "UPDATE plant_nodes SET label=? WHERE id=?",
                    (d.get('label'), node_id)
                )
            conn.commit()
        return jsonify({'ok': True})
    except Exception as e:
        logger.error(f"api_plant_patch_node error: {e}", exc_info=True)
        return jsonify({'error': 'Failed to update node'}), 500


@app.route('/api/plant/nodes/<int:node_id>', methods=['DELETE'])
def api_plant_delete_node(node_id):
    """Deactivate a plant node (soft-delete). Admin only."""
    if not session.get('user') or session['user'].get('role') != 'admin':
        return jsonify({'error': 'Admin only'}), 403
    try:
        with get_db_context() as conn:
            conn.execute(
                "UPDATE plant_nodes SET active=0 WHERE id=?", (node_id,)
            )
            conn.commit()
        return jsonify({'ok': True})
    except Exception as e:
        logger.error(f"api_plant_delete_node error: {e}", exc_info=True)
        return jsonify({'error': 'Failed to delete node'}), 500


# ---- Background Scheduler ----
def background_scheduler():
    """Background task scheduler with error recovery."""
    import schedule as sched
    
    shift_day = get_setting('shift_email_day', '06:00')
    shift_night = get_setting('shift_email_night', '07:00')
    
    def safe_task(func, task_name):
        """Wrap a scheduled task with error handling."""
        def wrapper():
            try:
                logger.info(f"Starting scheduled task: {task_name}")
                result = func()
                logger.info(f"Completed scheduled task: {task_name}")
                return result
            except Exception as e:
                logger.error(f"Scheduled task '{task_name}' failed: {e}", exc_info=True)
        return wrapper
    
    # Note: The actual email functions would need to be wrapped similarly
    # This is just the scheduler structure
    sched.every().monday.at("06:00").do(safe_task(generate_weekly_tasks, "weekly_task_generation"))
    sched.every().monday.at("06:30").do(safe_task(send_weekly_report, "weekly_report"))
    # sched.every().day.at(shift_day).do(safe_task(lambda: send_shift_breakdown_email("Day Shift"), "shift_email_day"))
    # sched.every().day.at(shift_night).do(safe_task(lambda: send_shift_breakdown_email("Night Shift"), "shift_email_night"))
    # sched.every().day.at("01:00").do(safe_task(lambda: send_monthly_archive() if date.today().day == 1 else None, "monthly_archive"))
    
    logger.info("Background scheduler started")
    while True:
        try:
            sched.run_pending()
            time.sleep(60)
        except Exception as e:
            logger.error(f"Scheduler loop error: {e}", exc_info=True)
            time.sleep(60)

# ---- Startup ----
if __name__ == '__main__':
    logger.info("=" * 60)
    logger.info("Mill Maintenance Terminal - Starting")
    logger.info("=" * 60)
    
    t = threading.Thread(target=background_scheduler, daemon=True)
    t.start()
    
    logger.info("Access on this machine: http://localhost:5000")
    logger.info("Access on network:      http://192.168.2.11:5000")
    logger.info("=" * 60)
    
    app.run(host='0.0.0.0', port=5000, debug=False)
